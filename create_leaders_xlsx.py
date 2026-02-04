import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Patent Research Leaders"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["Company", "Name", "Title", "Department Category", "LinkedIn URL", "Relevant Products/Tech Area"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# All company data
all_data = [
    # ============ NVIDIA ============
    # Product/Engineering Leaders
    {"Company": "NVIDIA", "Name": "Jan Kautz", "Title": "Vice President of Learning and Perception Research", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/jan-kautz-4668264", "Tech": "Computer vision, deep learning, visual perception, geometric vision, generative models, efficient deep learning, image denoising, super-resolution, object detection/recognition/classification, SLAM, optical flow"},
    {"Company": "NVIDIA", "Name": "Varsha Hedau", "Title": "Director, Computer Vision & Deep Learning", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/varsha-hedau-14ba4654/", "Tech": "Computer vision, deep learning, autonomous driving (L4), foundation models"},
    {"Company": "NVIDIA", "Name": "Sean Pieper", "Title": "Senior Director, Computer Vision", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/sean-pieper-5a234518/", "Tech": "Computer vision systems and applications"},
    {"Company": "NVIDIA", "Name": "Azita Martin", "Title": "Vice President & GM of AI for Retail, CPG, and QSR", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/azitamartin/", "Tech": "Retail AI, shrinkage detection, checkout systems, warehouse/distribution center AI, inventory management, computer vision for retail, edge computing for retail"},
    {"Company": "NVIDIA", "Name": "Deepu Talla", "Title": "Vice President & GM of Robotics & Edge Computing", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/deepu-talla-b1631bb", "Tech": "Edge AI, robotics, Jetson platform, embedded systems, generative AI on edge devices, autonomous systems"},
    {"Company": "NVIDIA", "Name": "Ming-Yu Liu", "Title": "Vice President of Research", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/mingyuliu/", "Tech": "GANs, generative AI, image synthesis, deep generative models, text2image, text2video, text2-3D, NVIDIA Canvas (GauGAN), NVIDIA Maxine"},
    {"Company": "NVIDIA", "Name": "Sanja Fidler", "Title": "Vice President of AI Research", "Category": "Product/Engineering", "LinkedIn": "https://ca.linkedin.com/in/sanja-fidler-2846a1a", "Tech": "3D computer vision, 3D reconstruction and synthesis, robotics simulation, multimodal representations, interactive labeling"},
    # IP/Legal
    {"Company": "NVIDIA", "Name": "Tim Teter", "Title": "Executive Vice President, General Counsel and Secretary", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/timothy-s-teter-0244604/", "Tech": "Top legal executive; prior experience litigating patent and technology matters at Cooley LLP; engineering background at Lockheed"},
    {"Company": "NVIDIA", "Name": "Rich Domingo", "Title": "Director of Intellectual Property", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/rich-domingo-a1b839b", "Tech": "IP portfolio management, patent strategy"},
    {"Company": "NVIDIA", "Name": "Kirk Bacon", "Title": "Senior Patent Counsel", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/baconkd/", "Tech": "Patent procurement, whitespace analysis, collaboration with R&D on inventive concepts"},
    # M&A/Corp Dev
    {"Company": "NVIDIA", "Name": "Vishal Bhagwati", "Title": "Head of Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/vishal-bhagwati-36956/", "Tech": "M&A strategy, acquisitions, strategic partnerships, corporate investments; prior CFO for AI/HPC at HPE"},
    {"Company": "NVIDIA", "Name": "Nico Caprez", "Title": "Corporate Development Manager", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/nicocaprez/", "Tech": "M&A execution, strategic investments, venture investments (e.g., MetAI digital twins)"},

    # ============ ALPHABET (GOOGLE) ============
    # Product/Engineering Leaders
    {"Company": "Alphabet (Google)", "Name": "Tomas Pfister", "Title": "Head of AI Research, Google Cloud", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/tomaspfister/", "Tech": "Deep learning, computer vision, facial micro-expression detection, human pose estimation, neural networks, Face ID technology foundations, LLMs, tabular data (TabNet), forecasting, anomaly detection"},
    {"Company": "Alphabet (Google)", "Name": "Pushmeet Kohli", "Title": "Vice President of Research, Google DeepMind", "Category": "Product/Engineering", "LinkedIn": "https://uk.linkedin.com/in/pushmeet-kohli-4838994", "Tech": "Computer vision, image segmentation, 3D reconstruction, text detection in images, AI safety (SynthID watermarking), probabilistic programming, deep learning verification"},
    {"Company": "Alphabet (Google)", "Name": "Rahul Sukthankar", "Title": "Vice President of Research, Google DeepMind", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/rahulsukthankar/", "Tech": "Computer vision, perception systems, AI powering Cloud, Nest, Pixel, Photos, Search, Waymo, YouTube; PhD in Robotics from CMU"},
    # IP/Legal
    {"Company": "Alphabet (Google)", "Name": "Michael Lee", "Title": "Director, Head of Patents", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/mleemichael/", "Tech": "Patent portfolio development, patent licensing and transactions, patent operations, data science for IP; former Senior Director IP at Cisco"},
    {"Company": "Alphabet (Google)", "Name": "Laura Sheridan", "Title": "Head of Patent Policy", "Category": "IP/Legal", "LinkedIn": "N/A - See Google patent policy team", "Tech": "Patent policy, patent examination process advocacy, AI and patents (testified before Congress), IP5 Industry engagement"},
    {"Company": "Alphabet (Google)", "Name": "Paul Leblond", "Title": "Senior Patent Counsel, Head of Devices and Services Patents", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/paul-leblond-73540111/", "Tech": "Devices and services IP, executive IP counseling, global patent portfolio development"},
    # M&A/Corp Dev
    {"Company": "Alphabet (Google)", "Name": "Don Harrison", "Title": "President, Global Partnerships and Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/don-harrison-a1b4271/", "Tech": "M&A strategy (led Waze, Motorola, Admob acquisitions), strategic partnerships, board member at SpaceX; former VP and Deputy General Counsel"},
    {"Company": "Alphabet (Google)", "Name": "Sandy Diep", "Title": "M&A Lead - DRS", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/sandydiep", "Tech": "M&A execution, deal structuring; prior experience at BitTorrent, Yahoo"},
]

# Write data
for row_idx, data in enumerate(all_data, 2):
    ws.cell(row=row_idx, column=1, value=data["Company"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=data["Name"]).border = thin_border
    ws.cell(row=row_idx, column=3, value=data["Title"]).border = thin_border
    ws.cell(row=row_idx, column=4, value=data["Category"]).border = thin_border
    ws.cell(row=row_idx, column=5, value=data["LinkedIn"]).border = thin_border
    ws.cell(row=row_idx, column=6, value=data["Tech"]).border = thin_border

    for col in range(1, 7):
        ws.cell(row=row_idx, column=col).alignment = cell_alignment

# Set column widths
column_widths = [18, 22, 55, 20, 55, 90]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze top row
ws.freeze_panes = "A2"

# Save
wb.save("/Users/charlieknott/Downloads/CMU SPRING /Ultron Internship/Patent_Research_Leaders.xlsx")
print("Excel file updated with NVIDIA and Google leaders!")
