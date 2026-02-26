#!/usr/bin/env python3
"""
build_litigation_dashboard.py

Reads index.html + Patlytics xlsx + Techson xlsx, merges all three data
sources into a single litigation-focused dashboard.

Output: litigation_dashboard.html
"""

import os, re, json, html as html_mod
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PATLYTICS_DIR = os.path.join(BASE_DIR, 'Patlytics')
TECHSON_FILE = os.path.join(BASE_DIR, 'Techson', 'patents_bundledxlsx.xlsx')
INPUT_HTML = os.path.join(BASE_DIR, 'base_dashboard.html')
OUTPUT_HTML = os.path.join(BASE_DIR, 'index.html')

# ──────────────────────────────────────────────
# Company name normalization
# ──────────────────────────────────────────────
COMPANY_NORM = {
    'google': 'Google', 'google (alphabet inc.)': 'Google',
    'google (alphabet)': 'Google', 'alphabet': 'Google', 'waymo': 'Google',
    'amazon': 'Amazon', 'amazon (aws)': 'Amazon',
    'apple': 'Apple', 'apple inc.': 'Apple',
    'microsoft': 'Microsoft', 'microsoft corporation': 'Microsoft',
    'nvidia': 'NVIDIA',
    'meta': 'Meta', 'meta (facebook)': 'Meta',
    'meta platforms (facebook)': 'Meta',
    'meta platforms inc. (facebook)': 'Meta',
    'samsung': 'Samsung', 'samsung electronics': 'Samsung',
    'hanwha vision (formerly samsung techwin)': 'Samsung',
    'tesla': 'Tesla', 'tesla inc.': 'Tesla',
    'openai': 'OpenAI',
    'qualcomm': 'Qualcomm',
    'softbank': 'SoftBank/ARM', 'arm': 'SoftBank/ARM',
    'xai': 'xAI',
}

TARGET_12 = ['Google', 'Amazon', 'Apple', 'Meta', 'Microsoft', 'NVIDIA',
             'Samsung', 'Tesla', 'OpenAI', 'Qualcomm', 'SoftBank/ARM', 'xAI']

def norm_company(name):
    """Normalize a company name to one of the 12 targets or return as-is."""
    key = name.strip().lower()
    return COMPANY_NORM.get(key, name.strip())

def norm_patent_id(pid):
    """Strip hyphens from patent IDs for cross-source matching."""
    return pid.replace('-', '').strip()

# ──────────────────────────────────────────────
# Load Patlytics data (10 xlsx files)
# ──────────────────────────────────────────────
def load_patlytics():
    by_patent = defaultdict(list)   # patent_id -> [{co, prod, score, docs, category}]
    by_product = defaultdict(list)  # (norm_co, prod) -> [{patent_id, score, category}]

    for fname in sorted(os.listdir(PATLYTICS_DIR)):
        if not fname.endswith('.xlsx') or fname.startswith('~'):
            continue
        m = re.search(r'Infringement_(.+)_2026', fname)
        category = m.group(1).replace('_', ' ') if m else fname

        wb = openpyxl.load_workbook(os.path.join(PATLYTICS_DIR, fname), data_only=True)
        ws = wb['Analysis']

        # Parse column headers (row 2, cols 5+)
        headers = []
        for c in range(5, ws.max_column + 1):
            h = ws.cell(2, c).value
            if not h:
                continue
            parts = h.split('\n')
            prod = parts[0].strip() if len(parts) > 0 else ''
            co = parts[1].strip() if len(parts) > 1 else ''
            docs_s = parts[2].strip() if len(parts) > 2 else ''
            docs_m = re.search(r'(\d+)', docs_s)
            docs = int(docs_m.group(1)) if docs_m else 0
            headers.append({'col': c, 'product': prod, 'company': co,
                            'company_norm': norm_company(co), 'docs': docs})

        # Parse patent rows (row 3+)
        for r in range(3, ws.max_row + 1):
            pid = ws.cell(r, 1).value
            if not pid:
                continue
            pid_n = norm_patent_id(pid)
            title = ws.cell(r, 2).value or ''

            for h in headers:
                score = ws.cell(r, h['col']).value
                if score is None:
                    continue
                score = float(score)
                if score <= 0:
                    continue
                entry = {
                    'co': h['company'], 'co_norm': h['company_norm'],
                    'prod': h['product'], 'score': round(score, 2),
                    'docs': h['docs'], 'category': category,
                }
                by_patent[pid_n].append(entry)
                by_product[(h['company_norm'], h['product'])].append({
                    'patent_id': pid_n, 'score': round(score, 2),
                    'category': category,
                })
        wb.close()

    return {'by_patent': dict(by_patent), 'by_product': dict(by_product)}


# ──────────────────────────────────────────────
# Load Techson data (bundled xlsx)
# ──────────────────────────────────────────────
def load_techson():
    wb = openpyxl.load_workbook(TECHSON_FILE, data_only=True)
    ws = wb['Patents']
    data = {}
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 3).value
        if not pid:
            continue
        pid_n = norm_patent_id(pid)

        target_cos_raw = ws.cell(r, 6).value or ''
        target_cos = [c.strip() for c in target_cos_raw.split('\n') if c.strip()]
        target_cos_norm = [norm_company(c) for c in target_cos]

        rev = ws.cell(r, 12).value
        exp = ws.cell(r, 13).value
        pri = ws.cell(r, 14).value

        def fmt_date(d):
            if isinstance(d, datetime):
                return d.strftime('%Y-%m-%d')
            return str(d) if d else '-'

        # Col 7: Relevant Products (newline-delimited list)
        prods_raw = ws.cell(r, 7).value or ''
        relevant_products = [p.strip() for p in prods_raw.split('\n') if p.strip()]

        data[pid_n] = {
            'bundle_no': ws.cell(r, 1).value or '',
            'bundle_title': ws.cell(r, 2).value or '',
            'title': ws.cell(r, 5).value or '',
            'quality': int(ws.cell(r, 8).value or 0),
            'tier1': int(ws.cell(r, 9).value or 0),
            'tier2': int(ws.cell(r, 10).value or 0),
            'tier3': int(ws.cell(r, 11).value or 0),
            'revenue': int(rev) if rev else 0,
            'expiration': fmt_date(exp),
            'priority_date': fmt_date(pri),
            'priority_desc': ws.cell(r, 15).value or '',
            'complexity': ws.cell(r, 16).value or '',
            'art_volume': ws.cell(r, 17).value or '',
            'diversity': ws.cell(r, 18).value or '',
            'status': ws.cell(r, 19).value or '',
            'target_cos': target_cos,
            'target_cos_norm': target_cos_norm,
            'relevant_products': relevant_products,
        }
    wb.close()
    return data


# ──────────────────────────────────────────────
# Format helpers
# ──────────────────────────────────────────────
def fmt_revenue(val):
    if val >= 1_000_000_000:
        return f'${val / 1_000_000_000:.1f}B'
    if val >= 1_000_000:
        return f'${val / 1_000_000:.0f}M'
    if val > 0:
        return f'${val / 1_000:.0f}K'
    return '-'

def score_class(s):
    if s >= 0.80: return 'score-cr'
    if s >= 0.60: return 'score-hi'
    if s >= 0.40: return 'score-md'
    return 'score-lo'

def quality_class(q):
    if q >= 7: return 'q-high'
    if q >= 4: return 'q-mid'
    return 'q-low'

def esc(text):
    return html_mod.escape(str(text)) if text else ''


# ──────────────────────────────────────────────
# Plain-English patent descriptions (for C-suite / non-technical readers)
# Key: normalized patent ID (no hyphens/slashes)
# ──────────────────────────────────────────────
PATENT_DESCRIPTIONS = {
    # 3D Reconstruction
    'WO2025193512A1': 'Covers a way to create a full 3D model of any object from a single photograph. The system looks at a flat photo, figures out what the object is, cuts it out from the background, and builds a three-dimensional version you can rotate and view from any angle.',
    'US10430922B2': 'Describes how to take one 2D image of an object and automatically generate a 3D model from it. The system figures out the object\'s shape and wraps the original photo\'s appearance onto the 3D version.',
    'US8861800B2': 'A fast method to create a 3D model of someone\'s face from just one photo. Works even when the photo is low quality or the person is not looking directly at the camera.',
    'US10755145B2': 'A system that takes a flat image of a face and transforms it into different angles and positions in 3D. It learns how faces rotate and reposition without needing multiple photos.',
    'US9916685B1': 'Figures out the depth (how far away each point is) of a face from a single photograph. Reconstructs the 3D shape of a face without needing any special camera equipment.',
    'US12322012B2': 'A way to correct perspective distortion in photos of scenes. Straightens and aligns images so objects appear as they would from a direct, head-on viewpoint.',

    # Object Detection
    'US10354159B2': 'Finds and identifies objects in images by looking at them at multiple sizes simultaneously. Faster and more accurate than checking one size at a time.',
    'US10354362B2': 'An improved way to detect objects in photos and videos by analyzing the image at several different scales. Draws boxes around each object it finds.',
    'US11954175B2': 'Detects objects in images using a layered approach that looks at both fine details and the bigger picture at the same time.',
    'US12026226B2': 'Teaches a computer to recognize new types of objects after seeing only a few examples. Uses the relationships between different object types to speed up learning.',
    'US12266156B2': 'Trains object detection systems even when the training images are missing some labels. Handles incomplete data without losing accuracy.',
    'US20240071029A1': 'A more flexible way to mark where objects are in an image. Instead of rigid reference points, it uses softer anchor points that adapt to each object\'s shape.',
    'US20240104761A1': 'Draws tilted or rotated bounding boxes around products in retail settings. Standard straight boxes miss the edges of angled items on shelves.',
    'US20240161469A1': 'A smarter way for AI systems to search back through an image when detecting objects. Remembers and reuses earlier findings to improve accuracy.',
    'US20240303979A1': 'Reduces false alarms in object detection. Helps systems avoid incorrectly identifying background clutter as real objects.',
    'US20240054775A1': 'Makes an object detection system trained in one setting work well in a completely different setting, without needing to retrain from scratch.',
    'US20250005881A1': 'Outlines irregularly shaped objects using complex polygon shapes instead of simple rectangles. Gives a much tighter fit around the actual object boundary.',
    'US20240320980A1': 'Automatically spots openly carried firearms in security camera footage. Can alert security personnel in real time.',

    # Face Recognition / Analysis
    'US10002286B1': 'A face recognition system that works even when photos are blurry, poorly lit, or low resolution. Identifies people in conditions where other systems fail.',
    'US10115004B2': 'Fills in missing or unclear facial details from a photo. Reconstructs what a face looks like based on the parts that are visible.',
    'US10121055B1': 'Pinpoints specific facial features (eyes, nose, mouth corners) in a photo. These landmarks are used as reference points for recognition and alignment.',
    'US12165068B2': 'Improves face recognition accuracy by better organizing the mathematical representations of different faces, making them easier to tell apart.',
    'US11900516B2': 'Recognizes a person\'s face even when they are looking sideways or at an angle. Generates what the face would look like from other positions to make a match.',
    'US9311564B2': 'Estimates how old someone is from a photo of their face. Works across different ages without needing to know the person\'s identity.',
    'US9336439B2': 'A camera system that captures clear images of someone\'s iris (the colored part of the eye) from several feet away, for identification without requiring close-up contact.',
    'US20140250523A1': 'Continuously verifies a user\'s identity while they use a device. Keeps checking that the same person is still present, not just at the initial login.',
    'US20200410210A1': 'A face recognition method that works regardless of which direction the person is facing. Matches faces across very different viewing angles.',
    'US7483569B2': 'A simpler and faster way to create image filters used for matching patterns in photos. These filters are a fundamental building block in many recognition systems.',
    'US9171226B2': 'Compares and matches images using techniques that capture both the texture and structure of what is in the photo. Useful as a foundation for more advanced matching systems.',

    # Neural Network Architecture
    'US12505663B2': 'Shrinks AI object detection models so they can run on smaller, less powerful devices like phones and cameras, without losing much accuracy.',
    'US11875557B2': 'A new type of AI network that uses polynomial math to process information more efficiently in the early layers, reducing the total computing work needed.',
    'US11989933B2': 'A variant of the polynomial AI network that applies its efficiency improvements in the later stages of processing, suited to different types of tasks.',
    'US12182707B2': 'Automatically discovers new, efficient AI network designs through controlled randomness, rather than requiring engineers to hand-design every connection.',
    'US20210034952A1': 'Builds AI networks that use small adjustments rather than full recalculations at each layer. Makes them faster and requires less computing power.',
    'US20240013032A1': 'AI networks that use simple on/off values instead of complex numbers. Makes them much faster and requires far less memory to run.',
    'US20240095524A1': 'Automatically designs the most efficient AI network structure for a given task, removing unnecessary parts to save processing power.',
    'US20240220801A1': 'Designs AI networks with fewer internal settings to tune, making them lighter and faster without sacrificing performance on the target task.',
    'US12131260B2': 'A technique for measuring how similar two items are by comparing their mathematical representations. Used as a core building block in recognition systems.',

    # Image Processing
    'US12136155B2': 'Generates realistic-looking synthetic images by separating and controlling different visual properties (shape, color, texture) independently.',
    'US11748853B2': 'Removes blur from photos when you do not know what caused the blur. Reverses the blurring process to recover a sharp image.',
    'US11854245B2': 'Trains AI image generators to produce better, more realistic images by ranking outputs from best to worst, instead of just scoring them individually.',
    'US20240054764A1': 'Creates varied training images by reshaping objects in existing photos. Helps AI models learn to recognize objects in different forms and poses.',
    'US20240320964A1': 'Creates new training images while keeping the identity of each object intact. Expands the training dataset without confusing the AI about what it is looking at.',

    # Few-Shot Learning
    'US12189714B2': 'Detects new types of objects after seeing only a few examples. Uses a network that dynamically adapts its understanding based on the limited samples available.',
    'US12190565B2': 'Reduces bias when training AI with very few examples. Selects training samples more carefully so the system does not lean toward common categories.',
    'US20230368038A1': 'An improved way to fine-tune AI models to recognize new objects from just a handful of labeled examples, getting better results than standard approaches.',
    'US20250181972A1': 'A fine-tuning method that looks at all test examples together to make better predictions when only a few training examples are available.',
    'US20240062532A1': 'Captures detailed visual features from images by pooling information from specific spatial locations. Helps the system distinguish between similar-looking objects.',

    # Retail: Product Detection
    'US12067527B2': 'Uses cameras to spot products that have been placed on the wrong shelf in a store. Compares what it sees against what should be there according to the store layout plan.',
    'US12437258B2': 'Identifies individual products on store shelves using camera images. Recognizes items by their packaging, labels, and position.',
    'US20240046621A1': 'Detects products in a retail setting, reads their text and barcodes, and matches them to a product database, all from camera footage.',
    'US20240355085A1': 'Matches products on shelves to their expected locations and identifies gaps or items that have shifted into the wrong spot.',
    'US20250182444A1': 'Recognizes products by looking at them from multiple camera angles. Combining views gives more reliable identification than a single shot.',

    # Retail: Checkout & Verification
    'US20250278988A1': 'Watches self-checkout stations to detect when items are not scanned or when someone tries to pay for a cheaper item than what they are actually taking.',

    # Action / Scene Understanding
    'US12327338B2': 'Improves the training data used to teach AI systems to recognize human actions (like walking, running, or waving) from body movement patterns.',
    'WO2024091472A1': 'Detects human actions in video by analyzing patterns of movement over time. Uses an efficient approach that does not require heavy computing resources.',

    # Verification / Search
    'US12217339B2': 'Verifies if two photos show the same object by testing multiple possible ways the object could have moved or rotated between shots.',
    'US12131497B2': 'A fast way to search through large collections of images to find a specific object. Filters out distractions the same way you focus on one voice in a crowded room.',
    'US11915463B2': 'Automatically builds and maintains a database of known objects from camera footage. Quickly identifies items it has seen before without manual data entry.',
    'WO2025175267A1': 'A two-way product verification system. Checks if a product matches its listing and if a listing matches the product, catching fakes from both directions.',
    'WO2025194159A1': 'Extends the two-way verification approach by using multiple camera views. Makes it harder to fool with carefully staged single-angle photos.',
    'US20250182363A1': 'The same multi-hypothesis verification method as US12217339B2, applied more broadly to new types of objects and product authentication scenarios.',
}


# ──────────────────────────────────────────────
# Patent category reorganization
# ──────────────────────────────────────────────

# Which patents to MOVE between categories:
#   - US12131260B2 (Cosine Embedding) → Neural Network Architecture
#   - US7483569B2, US9171226B2 (Other/Foundational) → Face Recognition / Analysis
#   - US20250278988A1 (Object Detection) → Retail: Checkout & Verification
# Which sections to DELETE: Cosine Embedding, Other / Foundational
# Which sections to SPLIT: Retail / Product AI → Product Detection + Checkout & Verification

PATENTS_TO_MOVE = {
    'US12131260B2': 'pat-neural-network-architecture',
    'US7483569B2': 'pat-face-recognition-analysis',
    'US9171226B2': 'pat-face-recognition-analysis',
    'US20250278988A1': 'pat-retail-checkout-verification',
}

SECTIONS_TO_DELETE = ['pat-cosine-embedding-similarity', 'pat-other-foundational']

# Retail split: patents that go to "Product Detection" vs "Checkout & Verification"
RETAIL_CHECKOUT_PATENTS = {'63/831,790', '63/794,781'}  # pre-filings (text match)
# US20250278988A1 comes from Object Detection via PATENTS_TO_MOVE
# Everything else in Retail / Product AI goes to Product Detection

NEW_RETAIL_PRODUCT_DESC = (
    'Patents for detecting, identifying, and matching products in retail settings — '
    'shelf scanning, planogram compliance, multiview recognition, and OCR-based scene reading.'
)
NEW_RETAIL_CHECKOUT_DESC = (
    'Patents for self-checkout verification, loss prevention, and pass-through checkout interfaces — '
    'shrinkage detection, barcode-free checkout, and mobile self-scan systems.'
)


def reorganize_patents(html_text):
    """Reorganize patent sections: merge orphan categories, split retail, move patents.

    Handles the fact that all patent rows may be on a single long line within
    each pat-section. Uses regex to parse and reconstruct sections.
    """
    # Parse all pat-section blocks from the HTML.
    # Each section: <div class="pat-section" id="xxx">...content...</div>
    # followed immediately by the next <div class="pat-section"> or end of patents page.
    # Because sections can be on one line, we use a non-greedy approach.

    # First, extract the patents page content
    page_start = html_text.find('id="page-patents"')
    if page_start < 0:
        print("  WARNING: Could not find patents page")
        return html_text

    # Find where the patent sections live (between <div class="sg">...</div> and </div> end of page)
    # The sections start after the stat grid
    first_section = html_text.find('<div class="pat-section"', page_start)
    if first_section < 0:
        print("  WARNING: Could not find any pat-section")
        return html_text

    # Find the end of all patent sections.
    # All sections are on one line. Find the last </div> that closes the last pat-section,
    # which is right before the newline or the closing </div> of the page.
    # Strategy: find the last occurrence of '</div>' before the page-closing '</div>\n'
    # by looking for the closing page tag pattern.
    close_page_re = re.search(r'</div>\s*\n\s*\n\s*<!--\s*COMPANY PAGES', html_text[first_section:])
    if close_page_re:
        # The sections end right before the page-closing </div>
        sections_end = first_section + close_page_re.start()
    else:
        # Fallback: find next page div
        next_page = html_text.find('<div class="page"', first_section + 1)
        if next_page > 0:
            sections_end = html_text.rfind('</div>', first_section, next_page)
        else:
            sections_end = len(html_text)

    sections_block = html_text[first_section:sections_end].strip()

    # Parse individual sections using regex
    # Each section: <div class="pat-section" id="ID">...followed by next section or end
    section_re = re.compile(r'<div class="pat-section" id="([^"]+)">')
    section_starts = [(m.start(), m.group(1)) for m in section_re.finditer(sections_block)]

    sections = {}  # id -> full HTML string
    for idx, (start, sec_id) in enumerate(section_starts):
        if idx + 1 < len(section_starts):
            end = section_starts[idx + 1][0]
        else:
            # Last section: find its closing </div> by counting depth
            end = len(sections_block)
        sections[sec_id] = sections_block[start:end]

    # Parse individual pat-rows from each section
    row_re = re.compile(r'<div class="pat-row">(.*?)</div>')

    def extract_rows(section_html):
        """Extract list of (full_row_html, patent_display_id) from a section."""
        rows = []
        for m in row_re.finditer(section_html):
            row_html = m.group(0)
            # Extract patent ID from pat-link or plain text
            link_m = re.search(r'class="pat-link"[^>]*>([^<]+)</a>', row_html)
            if link_m:
                pat_id = link_m.group(1).strip()
            else:
                # Pre-filing: <span>CMU Docket 2026-113</span> or <span>63/831,790</span>
                span_m = re.search(r'<span class="pat-doc"><span>([^<]+)</span>', row_html)
                pat_id = span_m.group(1).strip() if span_m else ''
            rows.append((row_html, pat_id))
        return rows

    def extract_header(section_html):
        """Extract the pat-cat header HTML."""
        m = re.search(r'<div class="pat-cat">.*?</div>', section_html)
        return m.group(0) if m else ''

    def extract_desc(section_html):
        """Extract the pat-cat-desc HTML."""
        m = re.search(r'<div class="pat-cat-desc">.*?</div>', section_html)
        return m.group(0) if m else ''

    def build_section(sec_id, title, count, desc, rows):
        """Build a pat-section HTML string."""
        parts = [f'<div class="pat-section" id="{sec_id}">']
        parts.append(f'<div class="pat-cat"><h3>{title}</h3><span class="pat-cnt">{count}</span></div>')
        parts.append(f'<div class="pat-cat-desc">{desc}</div>')
        for row_html, _ in rows:
            parts.append(row_html)
        parts.append('</div>')
        return ''.join(parts)

    # Step 1: Collect rows that need to move
    moved_rows = {}  # dest_section_id -> [row tuples]
    for dest in set(PATENTS_TO_MOVE.values()):
        moved_rows[dest] = []

    # Step 2: Process each section — remove moved-out rows
    processed_sections = {}
    for sec_id, sec_html in sections.items():
        rows = extract_rows(sec_html)
        kept = []
        for row_html, pat_id in rows:
            # Normalize pat_id for matching: remove slashes, spaces
            pat_clean = pat_id.replace('/', '').replace(' ', '')
            # Check if this patent needs to move
            moved = False
            for move_pat, move_dest in PATENTS_TO_MOVE.items():
                if pat_clean == move_pat.replace('/', '').replace(' ', ''):
                    moved_rows[move_dest].append((row_html, pat_id))
                    moved = True
                    break
            if not moved:
                kept.append((row_html, pat_id))
        processed_sections[sec_id] = kept

    # Step 3: Split Retail / Product AI
    retail_id = 'pat-retail-product-ai'
    if retail_id in processed_sections:
        retail_rows = processed_sections[retail_id]
        product_det_rows = []
        checkout_rows = []
        for row_html, pat_id in retail_rows:
            if pat_id in RETAIL_CHECKOUT_PATENTS:
                checkout_rows.append((row_html, pat_id))
            else:
                product_det_rows.append((row_html, pat_id))

        # Add any rows moved TO checkout from other sections
        checkout_rows.extend(moved_rows.get('pat-retail-checkout-verification', []))
        moved_rows.pop('pat-retail-checkout-verification', None)

        processed_sections['pat-retail-product-detection'] = product_det_rows
        processed_sections['pat-retail-checkout-verification'] = checkout_rows
        del processed_sections[retail_id]

    # Step 4: Add moved rows to their destination sections
    for dest, rows in moved_rows.items():
        if dest in processed_sections:
            processed_sections[dest].extend(rows)

    # Step 5: Remove deleted sections
    for sec_id in SECTIONS_TO_DELETE:
        processed_sections.pop(sec_id, None)

    # Step 6: Define the final ordering and metadata
    SECTION_META = {
        'pat-3d-reconstruction': ('3D Reconstruction',
            'Patents covering methods to build 3D models from flat images — including '
            'face reconstruction, depth estimation, and scene geometry from a single photo.'),
        'pat-object-detection': ('Object Detection',
            'Patents for identifying and locating objects in images and video — multiscale '
            'detection, bounding boxes, feature pyramids, and reducing false positives.'),
        'pat-face-recognition-analysis': ('Face Recognition / Analysis',
            'Patents for recognizing faces, estimating age, locating facial landmarks, '
            'iris-based identification, and foundational image matching — including methods '
            'robust to poor lighting, extreme angles, and degraded images.'),
        'pat-neural-network-architecture': ('Neural Network Architecture',
            'Patents for making AI models smaller, faster, and cheaper to run — compression, '
            'quantization, binary networks, automated architecture design, and cosine '
            'embedding techniques.'),
        'pat-image-processing': ('Image Processing',
            'Patents for generating, enhancing, and transforming images — photorealistic '
            'synthesis, deblurring, super-resolution, and data augmentation techniques.'),
        'pat-few-shot-learning': ('Few-Shot Learning',
            'Patents that let AI learn to recognize new objects from just a handful of '
            'examples — critical for applications where training data is scarce.'),
        'pat-retail-product-detection': ('Retail: Product Detection',
            NEW_RETAIL_PRODUCT_DESC),
        'pat-retail-checkout-verification': ('Retail: Checkout & Verification',
            NEW_RETAIL_CHECKOUT_DESC),
        'pat-action-scene-understanding': ('Action / Scene Understanding',
            'Patents for understanding what is happening in a scene — detecting human '
            'actions, skeleton tracking, and analyzing environment conditions.'),
        'pat-verification-search': ('Verification / Search',
            'Patents for verifying whether two objects are the same and rapidly searching '
            'large visual databases — product authentication and reverse lookup.'),
    }

    SECTION_ORDER = [
        'pat-3d-reconstruction',
        'pat-object-detection',
        'pat-face-recognition-analysis',
        'pat-neural-network-architecture',
        'pat-image-processing',
        'pat-few-shot-learning',
        'pat-retail-product-detection',
        'pat-retail-checkout-verification',
        'pat-action-scene-understanding',
        'pat-verification-search',
    ]

    # Step 7: Rebuild all sections in order
    new_sections = []
    total_patents = 0
    for sec_id in SECTION_ORDER:
        rows = processed_sections.get(sec_id, [])
        if not rows:
            continue
        title, desc = SECTION_META.get(sec_id, (sec_id, ''))
        new_sections.append(build_section(sec_id, title, len(rows), desc, rows))
        total_patents += len(rows)

    new_sections_html = ''.join(new_sections)

    # Step 8: Replace the old sections block with the new one
    # Also update the page-desc to reflect the category count
    # Preserve the newline before the page-closing </div>
    result = html_text[:first_section] + new_sections_html + '\n' + html_text[sections_end:]

    # Update category count in page description
    cat_count = len([s for s in SECTION_ORDER if s in processed_sections and processed_sections[s]])
    result = result.replace(
        '63 patents across 10 technology categories',
        f'{total_patents} patents across {cat_count} technology categories')

    print(f"  Reorganized: {total_patents} patents across {cat_count} categories")
    return result


# ──────────────────────────────────────────────
# CSS additions
# ──────────────────────────────────────────────
NEW_CSS = """
/* === LITIGATION DASHBOARD ADDITIONS === */
.ng-lit{background:rgba(91,163,217,.08)}
.src-badge{font-size:8px;font-weight:700;padding:1px 5px;border-radius:2px;text-transform:uppercase;letter-spacing:.3px;margin-right:4px;font-family:-apple-system,sans-serif;vertical-align:middle;display:inline-block}
.src-patlytics{background:rgba(155,89,182,.12);color:#8e44ad}
.src-techson{background:rgba(41,128,185,.12);color:#2980b9}
.src-dashboard{background:rgba(39,174,96,.12);color:#27ae60}

/* Patent row badges */
.pat-row{cursor:pointer;transition:background .15s}
.pat-row:hover{background:rgba(26,82,118,.03)}
.pat-badges{display:inline-flex;gap:6px;align-items:center;margin-left:auto;flex-shrink:0;font-family:-apple-system,sans-serif}
.pat-q{font-size:10px;font-weight:700;padding:2px 6px;border-radius:2px}
.q-high{background:rgba(39,174,96,.1);color:#27ae60}
.q-mid{background:rgba(214,137,16,.1);color:#d68910}
.q-low{background:rgba(192,57,43,.1);color:#c0392b}
.pat-rev{font-size:10px;color:var(--t2);font-weight:600}
.pat-top-inf{font-size:10px;color:var(--t2)}
.pat-top-inf strong{color:var(--t)}

/* Patent expandable detail */
.pat-detail{display:none;background:var(--bg);border:1px solid var(--b);border-radius:4px;margin:4px 0 12px;padding:14px 18px;font-size:12px;font-family:-apple-system,sans-serif;line-height:1.5}
.pat-detail.open{display:block}
.pat-detail-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
@media(max-width:768px){.pat-detail-grid{grid-template-columns:1fr}}
.pat-src-section{padding:10px 0}
.pat-src-section h4{font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--t3);margin-bottom:8px;display:flex;align-items:center;gap:6px}
.pat-meta-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:6px;margin-bottom:10px}
.pat-meta-item{font-size:11px}.pat-meta-item .lbl{color:var(--t3);font-size:9px;text-transform:uppercase}
.pat-score-tbl{width:100%;border-collapse:collapse;font-size:11px;margin-top:6px}
.pat-score-tbl th{text-align:left;padding:4px 8px;color:var(--t3);font-size:9px;text-transform:uppercase;border-bottom:1px solid var(--b)}
.pat-score-tbl td{padding:4px 8px;border-bottom:1px solid rgba(0,0,0,.03)}
.score-cr{color:var(--cr);font-weight:700}.score-hi{color:var(--hi);font-weight:600}.score-md{color:var(--md)}.score-lo{color:var(--t3)}

/* Patlytics badge on product cards */
.pi-pat-score{font-size:10px;font-weight:700;padding:2px 7px;border-radius:2px;background:rgba(155,89,182,.1);color:#8e44ad;font-family:-apple-system,sans-serif;margin-left:4px}
.pi-pat-none{font-size:10px;padding:2px 7px;border-radius:2px;background:rgba(139,145,154,.06);color:var(--t3);font-family:-apple-system,sans-serif;margin-left:4px}
/* Techson badge on product cards */
.pi-ts-badge{font-size:10px;font-weight:600;padding:2px 7px;border-radius:2px;background:rgba(41,128,185,.1);color:#2980b9;font-family:-apple-system,sans-serif;margin-left:4px}
/* Sort buttons on products toolbar */
.pi-sort-wrap{display:flex;align-items:center;gap:6px;margin-left:auto;font-size:11px;color:var(--t3);font-family:-apple-system,sans-serif}
.pi-sort-wrap label{font-size:10px;text-transform:uppercase;letter-spacing:.3px;font-weight:600;white-space:nowrap}
.pi-sort-btn{padding:3px 10px;border:1px solid var(--b);border-radius:3px;background:var(--s);cursor:pointer;font-size:11px;color:var(--t2);font-family:-apple-system,sans-serif;transition:all .15s}
.pi-sort-btn:hover{border-color:var(--a);color:var(--a)}
.pi-sort-btn.active{background:var(--a);color:#fff;border-color:var(--a)}

/* Infringement evidence section in product body */
.pi-evidence{margin:12px 0;padding:12px 14px;background:rgba(155,89,182,.03);border:1px solid rgba(155,89,182,.1);border-radius:4px}
.pi-evidence h4{font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:#8e44ad;margin-bottom:8px;display:flex;align-items:center;gap:6px}
.pi-ev-tbl{width:100%;border-collapse:collapse;font-size:11px}
.pi-ev-tbl th{text-align:left;padding:3px 8px;color:var(--t3);font-size:9px;text-transform:uppercase;border-bottom:1px solid rgba(155,89,182,.15)}
.pi-ev-tbl td{padding:3px 8px;border-bottom:1px solid rgba(0,0,0,.03)}

/* Litigation landing page */
.lit-page-desc{font-size:13px;color:var(--t2);margin-bottom:20px;line-height:1.5;font-style:italic;max-width:700px}
.lit-cards{display:flex;flex-direction:column;gap:12px}
.lit-card{background:var(--c);border:1px solid var(--b);border-radius:4px;box-shadow:var(--sh);overflow:hidden;transition:box-shadow .15s}
.lit-card:hover{box-shadow:0 2px 8px rgba(0,0,0,.06)}
.lit-head{display:flex;align-items:center;gap:12px;padding:14px 18px;cursor:pointer;user-select:none}
/* Override base co-mono to use rounded square everywhere */
.co-mono{border-radius:4px!important}
.nav-item .co-mono{vertical-align:middle;margin-right:4px;border-radius:3px!important}
h2 .co-mono{vertical-align:middle;margin-right:6px;border-radius:5px!important}
.net-co .co-mono{vertical-align:middle;margin-right:3px;border-radius:2px!important}
.lit-head .co-mono{width:28px;height:28px;display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:13px;border-radius:4px!important;flex-shrink:0;font-family:-apple-system,sans-serif}
.lit-head h3{font-size:15px;font-weight:700;flex:1;margin:0}
.lit-verdict{font-size:10px;font-weight:700;padding:3px 10px;border-radius:2px;text-transform:uppercase;letter-spacing:.3px;font-family:-apple-system,sans-serif}
.lv-high{background:rgba(192,57,43,.1);color:var(--cr)}.lv-moderate{background:rgba(214,137,16,.1);color:var(--hi)}.lv-low{background:rgba(139,145,154,.08);color:var(--t3)}
.lit-rev{font-size:12px;color:var(--t2);font-weight:600;font-family:-apple-system,sans-serif;margin-left:8px}
.lit-arrow{color:var(--t3);font-size:18px;transition:transform .15s}
.lit-card.open .lit-arrow{transform:rotate(90deg)}
.lit-body{display:none;padding:0 18px 16px;border-top:1px solid var(--b)}
.lit-card.open .lit-body{display:block}
.lit-src-row{padding:8px 0;border-bottom:1px solid rgba(0,0,0,.04);font-size:12px;color:var(--t2);line-height:1.6;font-family:-apple-system,sans-serif}
.lit-src-row:last-child{border-bottom:none}
.lit-src-row strong{color:var(--t)}
.lit-mini-tbl{width:100%;border-collapse:collapse;font-size:11px;margin-top:6px}
.lit-mini-tbl th{text-align:left;padding:3px 6px;color:var(--t3);font-size:9px;text-transform:uppercase}
.lit-mini-tbl td{padding:3px 6px}

/* Litigation summary on company tabs */
.lit-summary-card{background:var(--c);border:1px solid var(--b);border-radius:4px;box-shadow:var(--sh);margin-bottom:16px;padding:16px 18px}
.lit-summary-card h4{font-size:12px;font-weight:600;margin-bottom:10px;font-family:-apple-system,sans-serif;display:flex;align-items:center;gap:8px}
.lit-sum-stats{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:12px}
.lit-sum-stat{text-align:center}
.lit-sum-stat .lbl{font-size:9px;color:var(--t3);text-transform:uppercase;letter-spacing:.3px;font-family:-apple-system,sans-serif}
.lit-sum-stat .val{font-size:20px;font-weight:700;margin-top:2px}

/* Extended targets table */
.ext-targets{margin-top:28px}
.ext-targets h3{font-size:16px;margin-bottom:8px}
.ext-tbl{width:100%;border-collapse:collapse;font-size:12px;font-family:-apple-system,sans-serif}
.ext-tbl thead{position:sticky;top:0;z-index:2}
.ext-tbl th{text-align:left;padding:6px 10px;color:var(--t3);font-size:9px;text-transform:uppercase;letter-spacing:.3px;border-bottom:1px solid var(--b);background:var(--c)}
.ext-tbl td{padding:6px 10px;border-bottom:1px solid rgba(0,0,0,.04)}
.ext-tbl tr:hover{background:rgba(26,82,118,.02)}

/* Patent page filter bar */
.pat-filters{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:16px;font-family:-apple-system,sans-serif}
.pat-filter-sel{background:var(--bg);border:1px solid var(--b);color:var(--t);padding:5px 10px;border-radius:3px;font-size:11px;cursor:pointer;font-family:-apple-system,sans-serif}
.pat-filter-sel:focus{border-color:var(--a);outline:none}


/* Patent plain-English description */
.pat-desc-block{margin-top:12px;padding:10px 12px;background:rgba(26,82,118,.02);border-radius:4px}
.pat-desc-label{font-size:9px;text-transform:uppercase;letter-spacing:.5px;color:var(--t3);margin-bottom:4px;font-weight:600}
.pat-desc-text{font-size:12px;color:var(--t);line-height:1.6}

/* Category overview */
.cat-overview{margin-bottom:28px;font-family:-apple-system,sans-serif}
.cat-overview-hdr{font-size:14px;font-weight:600;margin-bottom:4px;color:var(--t)}
.cat-overview-sub{font-size:12px;color:var(--t3);margin-bottom:14px;line-height:1.5}
.cat-overview-tbl{width:100%;border-collapse:collapse;font-size:12px}
.cat-overview-tbl th{text-align:left;padding:6px 10px;font-size:9px;text-transform:uppercase;letter-spacing:.4px;color:var(--t3);border-bottom:2px solid var(--b);font-weight:600}
.cat-overview-tbl td{padding:7px 10px;border-bottom:1px solid rgba(0,0,0,.04);vertical-align:top;line-height:1.4}
.cat-overview-tbl tr:hover{background:rgba(26,82,118,.02)}
.cat-overview-tbl .cat-name{font-weight:600;color:var(--t);white-space:nowrap}
.cat-overview-tbl .cat-count{color:var(--t2);text-align:center;font-weight:600}
.cat-overview-tbl .cat-pl{font-size:11px;color:var(--t2)}
.cat-toggle{font-size:11px;color:var(--a);cursor:pointer;border:none;background:none;padding:0;text-decoration:underline;font-family:-apple-system,sans-serif;margin-bottom:12px;display:inline-block}
"""


# ──────────────────────────────────────────────
# Category overview for Patents page
# ──────────────────────────────────────────────
# Maps our category names → Patlytics source categories that were merged in
CATEGORY_PATLYTICS_MAP = {
    '3D Reconstruction': {
        'count': 6,
        'patlytics': '3D face modeling, reconstruction, and landmark pose normalization',
        'desc': 'Building 3D models from 2D images — face reconstruction, depth estimation, scene geometry from a single photo.',
    },
    'Object Detection': {
        'count': 12,
        'patlytics': 'General object detection models and training improvements',
        'desc': 'Identifying and locating objects in images/video — multiscale detection, bounding boxes, feature pyramids.',
    },
    'Face Recognition / Analysis': {
        'count': 11,
        'patlytics': 'Biometric face recognition and matching algorithms + Iris capture hardware',
        'desc': 'Recognizing faces, age estimation, facial landmarks, iris identification, and foundational image matching.',
    },
    'Neural Network Architecture': {
        'count': 9,
        'patlytics': 'Neural network architecture efficiency, compression, and training primitives',
        'desc': 'Making AI models smaller, faster, and cheaper — compression, quantization, binary networks, cosine embeddings.',
    },
    'Image Processing': {
        'count': 5,
        'patlytics': 'Included within neural network efficiency (Patlytics) — separated here for clarity',
        'desc': 'Generating, enhancing, and transforming images — photorealistic synthesis, deblurring, super-resolution.',
    },
    'Few-Shot Learning': {
        'count': 5,
        'patlytics': 'Few-shot learning and semantic conditioned recognition/detection',
        'desc': 'Learning to recognize new objects from just a handful of examples — critical when training data is scarce.',
    },
    'Retail: Product Detection': {
        'count': 6,
        'patlytics': 'Retail-oriented product detection, pose correction, and OCR association + Retail shelf product recognition and planogram compliance',
        'desc': 'Detecting and identifying products on shelves — scanning, planogram compliance, multiview recognition, OCR.',
    },
    'Retail: Checkout & Verification': {
        'count': 3,
        'patlytics': 'Retail checkout verification and anti-fraud monitoring',
        'desc': 'Self-checkout verification and loss prevention — shrinkage detection, barcode-free checkout, mobile self-scan.',
    },
    'Action / Scene Understanding': {
        'count': 3,
        'patlytics': 'Video surveillance threat detection and automated alerting (partial overlap)',
        'desc': 'Understanding what is happening in a scene — human action detection, skeleton tracking, environment analysis.',
    },
    'Verification / Search': {
        'count': 6,
        'patlytics': 'Included within general object detection (Patlytics) — separated here for IP clarity',
        'desc': 'Verifying whether two objects are the same and searching large visual databases — authentication, reverse lookup.',
    },
}


def build_category_overview():
    """Build the category overview + legend HTML for the top of the Patents page."""
    h = []
    h.append('<div class="cat-overview">')
    h.append('<h3 class="cat-overview-hdr">Category Overview</h3>')
    h.append('<p class="cat-overview-sub">Our 10 technology categories with corresponding Patlytics analysis groupings. '
             'Click any category below to jump to its patents.</p>')

    h.append('<table class="cat-overview-tbl"><thead><tr>')
    h.append('<th>Category</th><th style="text-align:center">Patents</th><th>Description</th><th>Patlytics Mapping</th>')
    h.append('</tr></thead><tbody>')

    for cat_name, info in CATEGORY_PATLYTICS_MAP.items():
        slug = cat_name.lower().replace(' / ', '-').replace(' ', '-').replace(':', '').replace('&', '')
        # Fix slug for the section IDs we actually use
        sec_id = {
            '3D Reconstruction': '3d-reconstruction',
            'Object Detection': 'object-detection',
            'Face Recognition / Analysis': 'face-recognition-analysis',
            'Neural Network Architecture': 'neural-network-architecture',
            'Image Processing': 'image-processing',
            'Few-Shot Learning': 'few-shot-learning',
            'Retail: Product Detection': 'retail-product-detection',
            'Retail: Checkout & Verification': 'retail-checkout-verification',
            'Action / Scene Understanding': 'action-scene-understanding',
            'Verification / Search': 'verification-search',
        }.get(cat_name, slug)

        h.append(f'<tr>')
        h.append(f'<td class="cat-name"><a href="#" onclick="document.getElementById(\'pat-{sec_id}\').scrollIntoView({{behavior:\'smooth\',block:\'start\'}});return false" '
                 f'style="color:var(--a);text-decoration:none">{esc(cat_name)}</a></td>')
        h.append(f'<td class="cat-count">{info["count"]}</td>')
        h.append(f'<td>{esc(info["desc"])}</td>')
        h.append(f'<td class="cat-pl"><span class="src-badge src-patlytics" style="font-size:7px">P</span> {esc(info["patlytics"])}</td>')
        h.append(f'</tr>')

    h.append('</tbody></table>')
    h.append('</div>')

    # Legend for Techson + Patlytics fields
    h.append('<div class="cat-overview" style="margin-bottom:20px">')
    h.append('<h3 class="cat-overview-hdr">Reading the Patent Data</h3>')
    h.append('<p class="cat-overview-sub">Each patent row shows inline badges. Click any patent to expand its full assessment.</p>')
    h.append('<table class="cat-overview-tbl"><thead><tr>')
    h.append('<th style="width:140px">Field</th><th style="width:80px">Source</th><th>What It Means</th>')
    h.append('</tr></thead><tbody>')

    legend_items = [
        ('Quality X/9', 'Techson', 'Patent strength score (1–9). Factors in claim breadth, prior art exposure, and enforceability. <strong>7+ = strong</strong>, 4–6 = moderate, &lt;4 = weak.'),
        ('$X.XB', 'Techson', 'Total annual revenue of products at risk of infringement by this patent, across all target companies.'),
        ('Company XX%', 'Patlytics', 'Highest claim-element match score for this patent against any product at that company. 100% = every claim element mapped to product evidence.'),
        ('Score (expanded)', 'Patlytics', 'Per-product infringement score (0–100%). Measures how many independent claim elements have documented evidence of use in that product.'),
        ('Evidence (expanded)', 'Patlytics', 'Number of source documents (patents, publications, product pages) Patlytics found supporting the infringement analysis.'),
        ('Revenue at Risk', 'Techson', 'Estimated revenue exposure for this specific patent, based on the target companies\' product revenue.'),
        ('Prior Art', 'Techson', 'Volume of existing prior art: LOW = less crowded (easier to enforce), HIGH = more prior art to navigate.'),
        ('Patent Status', 'Techson', 'Active = enforceable, Pending = application filed, Expired = no longer enforceable.'),
    ]

    for field, source, meaning in legend_items:
        badge_cls = 'src-techson' if source == 'Techson' else 'src-patlytics'
        h.append(f'<tr><td style="font-weight:600;white-space:nowrap">{esc(field)}</td>'
                 f'<td><span class="src-badge {badge_cls}" style="font-size:7px">{source[0]}</span> {source}</td>'
                 f'<td>{meaning}</td></tr>')

    h.append('</tbody></table>')
    h.append('</div>')

    return '\n'.join(h)


# ──────────────────────────────────────────────
# Build the Litigation Targets landing page
# ──────────────────────────────────────────────
COMPANY_COLORS = {
    'Google': '#4285F4', 'Amazon': '#FF9900', 'Apple': '#555555',
    'Meta': '#1877F2', 'Microsoft': '#00A4EF', 'NVIDIA': '#76B900',
    'Samsung': '#1428A0', 'Tesla': '#CC0000', 'OpenAI': '#412991',
    'Qualcomm': '#3253DC', 'SoftBank/ARM': '#ED1C24', 'xAI': '#000000',
}
COMPANY_LETTER = {
    'Google': 'G', 'Amazon': 'A', 'Apple': 'Ap', 'Meta': 'M',
    'Microsoft': 'Ms', 'NVIDIA': 'N', 'Samsung': 'Sa', 'Tesla': 'T',
    'OpenAI': 'O', 'Qualcomm': 'Q', 'SoftBank/ARM': 'S', 'xAI': 'x',
}

def co_icon(company, size=16):
    """Return HTML for a co-mono icon span at the given size."""
    color = COMPANY_COLORS.get(company, '#666')
    letter = COMPANY_LETTER.get(company, company[0])
    fs = max(8, size - 7)
    return (f'<span class="co-mono" style="background:{color};width:{size}px;height:{size}px;'
            f'font-size:{fs}px;line-height:{size}px">{letter}</span>')


def build_litigation_page(patlytics, techson):
    """Build HTML for the Litigation Targets landing page."""
    # Aggregate per-company stats
    co_stats = {}
    for co in TARGET_12:
        co_stats[co] = {
            'patlytics_scores': [],  # (product, score, patent_id)
            'techson_patents': 0,
            'techson_quality_sum': 0,
            'techson_revenue': 0,
        }

    # Patlytics aggregation
    for (co_norm, prod), entries in patlytics['by_product'].items():
        if co_norm in co_stats:
            best = max(entries, key=lambda e: e['score'])
            co_stats[co_norm]['patlytics_scores'].append(
                (prod, best['score'], best['patent_id'], best['category']))

    # Techson aggregation
    for pid, td in techson.items():
        for co_norm in set(td['target_cos_norm']):
            if co_norm in co_stats:
                co_stats[co_norm]['techson_patents'] += 1
                co_stats[co_norm]['techson_quality_sum'] += td['quality']
                co_stats[co_norm]['techson_revenue'] += td['revenue']

    # Sort each company's patlytics scores descending
    for co in co_stats:
        co_stats[co]['patlytics_scores'].sort(key=lambda x: -x[1])

    # Compute combined score for ranking
    def combined_score(co):
        cs = co_stats[co]
        best_pat = cs['patlytics_scores'][0][1] if cs['patlytics_scores'] else 0
        avg_q = (cs['techson_quality_sum'] / cs['techson_patents']
                 if cs['techson_patents'] else 0)
        return best_pat * 40 + avg_q * 3 + (cs['techson_revenue'] / 1e9) * 2

    ranked = sorted(TARGET_12, key=combined_score, reverse=True)

    # Stats
    total_rev = sum(td['revenue'] for td in techson.values())
    high_score_products = sum(
        1 for entries in patlytics['by_product'].values()
        if any(e['score'] >= 0.70 for e in entries)
    )
    high_q_patents = sum(1 for td in techson.values() if td['quality'] >= 7)

    html = []
    html.append(f'<div class="page" id="page-litigation">')
    html.append(f'<div class="flex-header"><h2>Litigation Targets</h2></div>')
    html.append(f'<p class="lit-page-desc">Cross-referenced infringement evidence from three independent analyses. '
                f'Each company card shows findings from Patlytics (claim-level scoring), '
                f'Techson (patent quality &amp; revenue risk), and our own product research.</p>')

    # Stat cards
    html.append(f'<div class="sg">')
    html.append(f'<div class="sc"><div class="l">Target Companies</div><div class="v v-a">{len(TARGET_12)}</div></div>')
    html.append(f'<div class="sc"><div class="l">Products &ge; 70% Match</div><div class="v v-cr">{high_score_products}</div></div>')
    html.append(f'<div class="sc"><div class="l">Portfolio Revenue Risk</div><div class="v v-hi">{fmt_revenue(total_rev)}</div></div>')
    html.append(f'<div class="sc"><div class="l">High-Quality Patents (7+)</div><div class="v v-gn">{high_q_patents}</div></div>')
    html.append(f'</div>')

    # Company cards
    html.append(f'<div class="lit-cards">')
    for co in ranked:
        cs = co_stats[co]
        best_score = cs['patlytics_scores'][0][1] if cs['patlytics_scores'] else 0
        avg_q = (cs['techson_quality_sum'] / cs['techson_patents']
                 if cs['techson_patents'] else 0)

        if best_score >= 0.80 or cs['techson_revenue'] >= 5e9:
            verdict, vc = 'HIGH RISK', 'lv-high'
        elif best_score >= 0.50 or cs['techson_revenue'] >= 1e9:
            verdict, vc = 'MODERATE', 'lv-moderate'
        else:
            verdict, vc = 'LOW', 'lv-low'

        color = COMPANY_COLORS.get(co, '#666')
        letter = COMPANY_LETTER.get(co, co[0])

        html.append(f'<div class="lit-card" data-co="{esc(co)}">')
        html.append(f'<div class="lit-head" onclick="toggleLit(this.parentElement)">')
        html.append(f'<span class="co-mono" style="background:{color}">{letter}</span>')
        html.append(f'<h3>{esc(co)}</h3>')
        html.append(f'<span class="lit-verdict {vc}">{verdict}</span>')
        if cs['techson_revenue']:
            html.append(f'<span class="lit-rev">{fmt_revenue(cs["techson_revenue"])} exposure</span>')
        html.append(f'<span class="lit-arrow">&#x203A;</span>')
        html.append(f'</div>')

        # Body
        html.append(f'<div class="lit-body">')

        # Patlytics row
        if cs['patlytics_scores']:
            top3 = cs['patlytics_scores'][:5]
            html.append(f'<div class="lit-src-row">')
            html.append(f'<span class="src-badge src-patlytics">Patlytics</span> ')
            parts = []
            for prod, sc, pat, cat in top3:
                cls = score_class(sc)
                parts.append(f'{esc(prod)}: <strong class="{cls}">{sc:.0%}</strong>')
            html.append(' &bull; '.join(parts))
            html.append(f'<br><span style="font-size:10px;color:var(--t3)">{len(cs["patlytics_scores"])} products analyzed across {len(set(e[2] for e in cs["patlytics_scores"]))} patents</span>')
            html.append(f'</div>')
        else:
            html.append(f'<div class="lit-src-row"><span class="src-badge src-patlytics">Patlytics</span> No products scored for {esc(co)}</div>')

        # Techson row
        html.append(f'<div class="lit-src-row">')
        html.append(f'<span class="src-badge src-techson">Techson</span> ')
        if cs['techson_patents']:
            html.append(f'<strong>{cs["techson_patents"]}</strong> patents target {esc(co)} &bull; ')
            html.append(f'Avg quality: <strong>{avg_q:.1f}</strong>/9 &bull; ')
            html.append(f'Revenue risk: <strong>{fmt_revenue(cs["techson_revenue"])}</strong>')
        else:
            html.append(f'No patents targeting {esc(co)} in Techson analysis')
        html.append(f'</div>')

        # Our Research row (placeholder — will be populated from existing HTML data)
        html.append(f'<div class="lit-src-row">')
        html.append(f'<span class="src-badge src-dashboard">Our Research</span> ')
        html.append(f'See <a href="#" onclick="goToCompany(\'{esc(co)}\');return false" style="color:var(--a)">{esc(co)} tab</a> for product mappings, contacts &amp; conflict ratings')
        html.append(f'</div>')

        html.append(f'</div>')  # lit-body
        html.append(f'</div>')  # lit-card

    html.append(f'</div>')  # lit-cards

    # Extended targets (non-12 companies with high scores)
    ext_targets = []
    for (co_norm, prod), entries in patlytics['by_product'].items():
        if co_norm in TARGET_12:
            continue
        best = max(entries, key=lambda e: e['score'])
        ext_targets.append({
            'company': co_norm, 'product': prod,
            'score': best['score'], 'patent_id': best['patent_id'],
            'category': best['category'], 'patent_count': len(entries),
        })
    ext_targets.sort(key=lambda x: -x['score'])

    if ext_targets:
        html.append(f'<div class="ext-targets">')
        html.append(f'<h3>Extended Targets (Beyond Core 12)</h3>')
        html.append(f'<p class="page-desc" style="margin-bottom:12px">Companies identified by Patlytics with significant infringement scores but outside the current 12-company focus.</p>')
        html.append(f'<table class="ext-tbl"><thead><tr>')
        html.append(f'<th>Company</th><th>Product</th><th>Best Score</th><th>Patents</th><th>Category</th>')
        html.append(f'</tr></thead><tbody>')
        seen = set()
        for et in ext_targets[:30]:
            key = (et['company'], et['product'])
            if key in seen:
                continue
            seen.add(key)
            cls = score_class(et['score'])
            html.append(f'<tr><td><strong>{esc(et["company"])}</strong></td>'
                         f'<td>{esc(et["product"])}</td>'
                         f'<td class="{cls}">{et["score"]:.0%}</td>'
                         f'<td>{et["patent_count"]}</td>'
                         f'<td style="font-size:10px;color:var(--t3)">{esc(et["category"])}</td></tr>')
        html.append(f'</tbody></table>')
        html.append(f'</div>')

    html.append(f'</div>')  # page-litigation
    return '\n'.join(html)


# ──────────────────────────────────────────────
# Enhance patent rows
# ──────────────────────────────────────────────
def enhance_patent_rows(html_text, patlytics, techson):
    """Inject Techson + Patlytics badges into patent rows and add expandable details.
    NOTE: All patent rows may be on a single long line, so we use regex on the full string."""

    def build_detail(pid_n, ts, pl_entries):
        """Build expandable detail panel HTML for a patent."""
        d = [f'<div class="pat-detail" id="pd-{pid_n}">',
             f'<div class="pat-detail-grid">']

        # Left: Techson
        d.append('<div class="pat-src-section">')
        d.append('<h4><span class="src-badge src-techson">Techson</span> Patent Assessment</h4>')
        if ts:
            d.append('<div class="pat-meta-grid">')
            d.append(f'<div class="pat-meta-item"><div class="lbl">Quality Score</div><div><strong>{ts.get("quality","?")}</strong>/9</div></div>')
            d.append(f'<div class="pat-meta-item"><div class="lbl">Revenue at Risk</div><div><strong>{fmt_revenue(ts.get("revenue",0))}</strong></div></div>')
            d.append(f'<div class="pat-meta-item"><div class="lbl">Patent Status</div><div>{esc(ts.get("status","?"))}</div></div>')
            d.append(f'<div class="pat-meta-item"><div class="lbl">Expiration</div><div>{esc(ts.get("expiration","-"))}</div></div>')
            d.append(f'<div class="pat-meta-item"><div class="lbl">Prior Art</div><div>{esc(ts.get("art_volume","?"))}</div></div>')
            d.append('</div>')
            target_12_here = sorted(set(c for c in ts.get('target_cos_norm', []) if c in TARGET_12))
            if target_12_here:
                d.append(f'<div style="font-size:11px;margin-top:4px"><span style="color:var(--t3)">Targets:</span> {", ".join(esc(c) for c in target_12_here)}</div>')
            others = [c for c in ts.get('target_cos', []) if norm_company(c) not in TARGET_12]
            if others:
                d.append(f'<div style="font-size:10px;color:var(--t3);margin-top:2px">+ {", ".join(esc(c) for c in others[:8])}{"..." if len(others) > 8 else ""}</div>')
        else:
            d.append('<div style="font-size:11px;color:var(--t3)">Not in Techson analysis (WIPO/pre-filing patent)</div>')

        # Plain-English description (fills blank space on left side)
        desc = PATENT_DESCRIPTIONS.get(pid_n, '')
        if desc:
            d.append(f'<div class="pat-desc-block"><div class="pat-desc-label">What This Patent Does</div><div class="pat-desc-text">{esc(desc)}</div></div>')
        d.append('</div>')

        # Right: Patlytics
        d.append('<div class="pat-src-section">')
        d.append('<h4><span class="src-badge src-patlytics">Patlytics</span> Infringement Scores</h4>')
        if pl_entries:
            sorted_e = sorted(pl_entries, key=lambda e: -e['score'])
            visible_count = 12
            d.append(f'<table class="pat-score-tbl" id="pst-{pid_n}"><thead><tr><th>Company</th><th>Product</th><th>Score</th><th title="Number of source documents supporting the infringement analysis">Evidence</th></tr></thead><tbody>')
            for i, e in enumerate(sorted_e):
                cls = score_class(e['score'])
                hidden = ' style="display:none" class="pst-extra"' if i >= visible_count else ''
                d.append(f'<tr{hidden}><td>{esc(e["co_norm"])}</td><td>{esc(e["prod"])}</td>'
                         f'<td class="{cls}">{e["score"]:.0%}</td><td>{e["docs"]} docs</td></tr>')
            d.append('</tbody></table>')
            if len(sorted_e) > visible_count:
                d.append(f'<div style="margin-top:4px"><a href="#" onclick="togglePatScoreRows(\'{pid_n}\',this);return false" '
                         f'style="font-size:10px;color:var(--a)">Show {len(sorted_e)-visible_count} more</a></div>')
        else:
            d.append('<div style="font-size:11px;color:var(--t3)">No Patlytics scores available</div>')
        d.append('</div>')

        d.append('</div></div>')  # close grid + detail
        return ''.join(d)

    # Process each pat-row using regex on the full HTML string
    # Pattern: <div class="pat-row">...<a ...class="pat-link">PATENT_ID</a>...</div>
    # IMPORTANT: The middle group must NOT cross </div> boundaries — otherwise it
    # can span from a pre-filing row (no pat-link) into the next section's row.
    # We use a negative lookahead (?:(?!</div>).)* to ensure we stay within one row.
    pat_row_re = re.compile(
        r'(<div class="pat-row">)((?:(?!</div>).)*?class="pat-link"[^>]*>)([^<]+)(</a>(?:(?!</div>).)*?)(</div>)',
        re.DOTALL)

    def replace_row(m):
        prefix = m.group(1)      # <div class="pat-row">
        mid1 = m.group(2)        # ...class="pat-link"...>
        pat_display = m.group(3) # patent ID text
        mid2 = m.group(4)        # </a>...remaining spans
        close = m.group(5)       # </div>

        pid_n = norm_patent_id(pat_display.replace('/', ''))
        ts = techson.get(pid_n, {})
        pl_entries = patlytics['by_patent'].get(pid_n, [])

        # Build badges
        badges = []
        if ts:
            q = ts.get('quality', 0)
            badges.append(f'<span class="pat-q {quality_class(q)}" title="Techson Quality Score">Quality {q}/9</span>')
            rev = ts.get('revenue', 0)
            if rev:
                badges.append(f'<span class="pat-rev" title="Techson Revenue Risk">{fmt_revenue(rev)}</span>')
        if pl_entries:
            best = max(pl_entries, key=lambda e: e['score'])
            cls = score_class(best['score'])
            badges.append(f'<span class="pat-top-inf"><span class="src-badge src-patlytics">P</span> {esc(best["co_norm"])} <strong class="{cls}">{best["score"]:.0%}</strong></span>')

        badge_html = f'<span class="pat-badges">{"".join(badges)}</span>' if badges else ''

        # Add onclick and badges
        new_prefix = f'<div class="pat-row" onclick="togglePatDetail(\'{pid_n}\')">'
        row_html = f'{new_prefix}{mid1}{pat_display}{mid2}{badge_html}{close}'

        # Add expandable detail after the row
        detail_html = build_detail(pid_n, ts, pl_entries)
        return row_html + detail_html

    result = pat_row_re.sub(replace_row, html_text)
    return result


# ──────────────────────────────────────────────
# Techson → Dashboard product mapping
# ──────────────────────────────────────────────
# Maps (company, dashboard_product_name) → True when Techson's "Relevant Products"
# reference the same product (even if names differ slightly).
# Derived by matching Techson Col 7 product names to our 12 target companies,
# then fuzzy-matching against dashboard product names.

def build_techson_product_set(techson):
    """Build set of dashboard products that Techson also identifies.
    Returns: { (company, dash_product_name): techson_product_name }"""

    # Company detection patterns for Techson product names
    CO_PAT = {
        'Google': re.compile(r'\bgoogle\b|\balphabet\b|\bwaymo\b|\byoutube\b', re.I),
        'Amazon': re.compile(r'\bamazon\b|\baws\b|\bring\b(?!\s*(?:buffer|topology|network))', re.I),
        'Apple': re.compile(r'\bapple\b|\bface\s*id\b|\bcore\s*ml\b|\bios\s+photos\b|\biphone\b|\bipad\b', re.I),
        'Microsoft': re.compile(r'\bmicrosoft\b|\bazure\b|\bwindows\s*hello\b|\bkinect\b|\bbing\b', re.I),
        'Meta': re.compile(r'\bmeta\s+(?:ai|quest)\b|\bfacebook\b|\binstagram\b|\bpytorch\b', re.I),
        'NVIDIA': re.compile(r'\bnvidia\b|\bjetson\b|\btensorrt\b', re.I),
        'Tesla': re.compile(r'\btesla\b|\bautopilot\b', re.I),
        'Samsung': re.compile(r'\bsamsung\b|\bgalaxy\b|\bbixby\b', re.I),
        'OpenAI': re.compile(r'\bopenai\b|\bgpt-?4\b|\bdall.?e\b|\bchatgpt\b', re.I),
        'Qualcomm': re.compile(r'\bqualcomm\b|\bsnapdragon\b', re.I),
        'SoftBank/ARM': re.compile(r'\bsoftbank\b|\barm\s+ethos\b|\barm\s+nn\b', re.I),
    }

    # Collect all Techson products mapped to our companies
    ts_by_co = {}  # company -> set of product names
    for pid, td in techson.items():
        for prod in td.get('relevant_products', []):
            for co, pat in CO_PAT.items():
                if pat.search(prod):
                    ts_by_co.setdefault(co, set()).add(prod)

    # Manual overrides: (company, dashboard_product) -> True
    # These handle cases where names differ significantly but refer to the same product
    MANUAL_MATCHES = {
        ('Amazon', 'Just Walk Out Technology'): 'Amazon Go Store Automated Checkout System',
        ('Amazon', 'Rekognition'): 'Amazon Rekognition',
        ('Amazon', 'Ring Smart Alerts'): 'Ring Video Doorbell',
        ('Amazon', 'SageMaker Neo / JumpStart'): 'Amazon SageMaker',
        ('Amazon', 'Panorama / Bedrock'): 'AWS DeepLens',
        ('Google', 'Cloud Vision API / Vertex AI'): 'Google Cloud Vision AI',
        ('Google', 'Google Lens'): 'Google Lens',
        ('Google', 'Photos Face Grouping'): 'Google Photos',
        ('Google', 'Pixel Face Unlock / Nest Cam'): 'Pixel Face Unlock',
        ('Google', 'Waymo Perception Stack'): 'Waymo Autonomous Vehicles',
        ('Apple', 'Face ID'): 'Face ID',
        ('Apple', 'Core ML Compression'): 'Core ML (Vision Framework)',
        ('Apple', 'Apple Neural Engine'): 'Apple Neural Engine',
        ('Apple', 'Portrait Mode / Photos Face Recognition'): 'Apple Photos',
        ('Apple', 'SHARP / Vision Framework'): 'Core ML (Vision Framework)',
        ('Apple', 'Siri Visual Intelligence / Image Understanding'): 'Apple Photos (Visual Lookup)',
        ('Apple', 'ARKit / Object Capture'): 'iOS Photos App',
        ('Microsoft', 'Azure AI Face API'): 'Azure Face API',
        ('Microsoft', 'Azure AI Vision / Custom Vision'): 'Azure Custom Vision',
        ('Microsoft', 'Windows Hello Face'): 'Windows Hello',
        ('Microsoft', 'Surface / Kinect'): 'Azure Kinect DK',
        ('Microsoft', 'HoloLens 2'): 'Microsoft Mesh',
        ('Microsoft', 'Copilot Vision / Designer'): 'Microsoft Designer',
        ('Meta', 'Codec Avatars / Quest Face Tracking'): 'Meta Quest 3 (facial mapping for avatars in VR)',
        ('Meta', 'SAM / Detectron2 / 3D Gen / DINOv2'): 'Meta AI Research Object Detection Models',
        ('Meta', 'SAM / Detectron2 / DensePose'): 'Meta AI Research Object Detection Models',
        ('Meta', 'FBGEMM / FBNet / Distillation'): 'PyTorch',
        ('Meta', 'Quest Hardware'): 'Meta Quest (3D Model Import/Creation Tools)',
        ('Meta', 'Content Understanding'): 'Facebook Photo Tagging and AI Recognition',
        ('NVIDIA', 'Retail Microservices / DeepStream'): 'NVIDIA DeepStream SDK',
        ('NVIDIA', 'TensorRT / TAO Toolkit'): 'NVIDIA TAO Toolkit',
        ('NVIDIA', 'AI Blueprints (Video Analytics Agents)'): 'NVIDIA Metropolis',
        ('NVIDIA', 'Maxine Face AI / StyleGAN'): 'NVIDIA Maxine',
        ('NVIDIA', 'Jetson Orin'): 'Jetson Nano Developer Kit',
        ('NVIDIA', 'Clara Holoscan'): 'NVIDIA Clara Imaging',
        ('NVIDIA', 'StyleGAN / GauGAN / NeMo'): 'NVIDIA NeMo',
        ('NVIDIA', 'GPU Compute Platform'): 'NVIDIA DGX Platform',
        ('Tesla', 'Tesla Vision / Occupancy Networks / HydraNet'): 'Tesla Autopilot (Vision System)',
        ('Tesla', 'FSD Computer HW4'): 'Tesla FSD (Full Self-Driving Computer)',
        ('Tesla', 'FSD End-to-End NN'): 'Tesla FSD (Full Self-Driving) Computer Vision System',
        ('Tesla', 'Vision Park Assist (Camera-Only 3D)'): 'Tesla Autopilot',
        ('Samsung', 'Bixby Vision / Circle to Search'): 'Samsung Galaxy S23 (Bixby Vision)',
        ('Samsung', 'Iris Scanner / Face Recognition / Galaxy AI Photo Assist'): 'Samsung Galaxy S23 (Bixby Vision)',
        ('OpenAI', 'GPT-4o Native Image Gen (gpt-image-1)'): 'GPT-4 Vision',
        ('OpenAI', 'GPT-4o Vision / DALL-E / Shap-E'): 'OpenAI Vision-enabled GPT-4',
        ('OpenAI', 'CLIP / Shap-E'): 'OpenAI CLIP',
        ('OpenAI', 'ChatGPT Product'): 'ChatGPT',
        ('OpenAI', 'API (Vision + Generation endpoints)'): 'OpenAI API',
        ('Qualcomm', 'Snapdragon X2 (AI PC NPU)'): 'Qualcomm AI Engine (Snapdragon platforms)',
        ('Qualcomm', 'Hexagon NPU / Snapdragon 8 Elite Gen 5'): 'Qualcomm Snapdragon Neural Processing Engine',
        ('Qualcomm', 'Spectra ISP / Face Detection'): 'Qualcomm Snapdragon Neural Processing Engine',
        ('SoftBank/ARM', 'Ethos NPU / ARM NN / Compute Library'): 'Arm Ethos-U NPU',
        ('SoftBank/ARM', 'Ethos-U85 Edge AI Platform'): 'Arm Ethos-U NPU',
    }

    return MANUAL_MATCHES


# ──────────────────────────────────────────────
# Enhance product cards
# ──────────────────────────────────────────────
def build_product_patlytics_lookup(patlytics):
    """Build lookup: (norm_company, product_name) -> best_score + patent list.
    Also try matching dashboard product names to Patlytics product names."""
    lookup = {}
    for (co_norm, prod), entries in patlytics['by_product'].items():
        best = max(entries, key=lambda e: e['score'])
        sorted_e = sorted(entries, key=lambda e: -e['score'])
        lookup[(co_norm, prod)] = {
            'best_score': best['score'],
            'best_patent': best['patent_id'],
            'entries': sorted_e[:8],
        }
    return lookup


def _tokenize(s):
    """Extract meaningful words from a product name (lowercase)."""
    return set(re.findall(r'[a-z][a-z0-9]+', s.lower()))

def find_patlytics_for_product(dash_prod, dash_co, pl_lookup):
    """Try to find Patlytics data for a dashboard product.
    Uses: 1) exact match, 2) substring match, 3) word overlap match."""
    # 1. Exact match
    key = (dash_co, dash_prod)
    if key in pl_lookup:
        return pl_lookup[key]

    dash_lower = dash_prod.lower()
    candidates = [(co, prod, data) for (co, prod), data in pl_lookup.items()
                  if co == dash_co]

    # 2. Substring match
    best_match = None
    best_score = 0
    for co, prod, data in candidates:
        prod_lower = prod.lower()
        if dash_lower in prod_lower or prod_lower in dash_lower:
            if data['best_score'] > best_score:
                best_match = data
                best_score = data['best_score']
    if best_match:
        return best_match

    # 3. Word overlap: dashboard products often use "X / Y" format or descriptive names.
    #    Match if a Patlytics product shares 2+ significant words with dashboard product.
    dash_tokens = _tokenize(dash_prod)
    # Remove very common/short words
    stop_words = {'the', 'and', 'for', 'with', 'from', 'pro', 'gen', 'new', 'via'}
    dash_tokens -= stop_words

    if len(dash_tokens) < 2:
        return None

    for co, prod, data in candidates:
        pl_tokens = _tokenize(prod) - stop_words
        overlap = dash_tokens & pl_tokens
        # Need at least 2 word overlap, or 1 word if it's a distinctive product name (4+ chars)
        distinctive = [w for w in overlap if len(w) >= 4]
        if len(overlap) >= 2 or (len(distinctive) >= 1 and len(overlap) >= 1):
            if data['best_score'] > best_score:
                best_match = data
                best_score = data['best_score']

    return best_match


def enhance_product_cards(html_text, patlytics, techson):
    """Add Patlytics infringement score and Techson overlap badge to product card headers.
    Also adds data-patlytics and data-techson attributes for sorting.
    Uses per-line processing: each pi card starts on its own line, and each card's
    sub-elements (pi-title, pi-body, pi-contacts) are on separate lines within a
    15-line block."""
    pl_lookup = build_product_patlytics_lookup(patlytics)
    ts_product_map = build_techson_product_set(techson)

    lines = html_text.split('\n')
    new_lines = []
    current_card_data = None  # Track which card we're in
    current_card_techson = False  # Track Techson overlap

    for i, line in enumerate(lines):
        # Detect start of a product card
        pi_match = re.search(r'<div class="pi"[^>]*data-prod="([^"]*)"[^>]*data-company="([^"]*)"', line)
        if pi_match:
            prod_name = html_mod.unescape(pi_match.group(1))
            company = html_mod.unescape(pi_match.group(2))
            current_card_data = find_patlytics_for_product(prod_name, company, pl_lookup)
            current_card_techson = (company, prod_name) in ts_product_map

            # Add data attributes for sorting
            pat_score = current_card_data['best_score'] if current_card_data else 0
            ts_val = 1 if current_card_techson else 0
            line = line.replace(
                f'data-prod="{pi_match.group(1)}"',
                f'data-prod="{pi_match.group(1)}" data-patlytics="{pat_score:.2f}" data-techson="{ts_val}"'
            )

        # Inject badges on contact-badge line (within current card)
        if (current_card_data is not None or current_card_techson) and 'contact-badge' in line and 'pi-title' in line:
            badges = ''
            if current_card_data:
                score_pct = f'{current_card_data["best_score"]:.0%}'
                badges += f' <span class="pi-pat-score" title="Patlytics best infringement score">{score_pct}</span>'
            if current_card_techson:
                badges += ' <span class="pi-ts-badge" title="Techson also identifies this product">Techson</span>'
            if badges:
                line = line.replace('</span></div>', f'</span>{badges}</div>', 1)

        # Inject evidence section before pi-contacts (within current card)
        if current_card_data is not None and 'class="pi-contacts"' in line:
            if current_card_data['entries']:
                ev = ['<div class="pi-evidence">']
                ev.append('<h4><span class="src-badge src-patlytics">Patlytics</span> Infringement Evidence</h4>')
                ev.append('<table class="pi-ev-tbl"><thead><tr><th>Patent</th><th>Score</th><th>Category</th></tr></thead><tbody>')
                for e in current_card_data['entries'][:6]:
                    cls = score_class(e['score'])
                    pat = e['patent_id']
                    ev.append(f'<tr><td style="font-size:10px"><a href="#" onclick="goToPatent(\'{pat}\');return false" '
                              f'style="color:var(--a);text-decoration:none">{pat}</a></td>'
                              f'<td class="{cls}">{e["score"]:.0%}</td>'
                              f'<td style="font-size:10px;color:var(--t3)">{esc(e["category"])}</td></tr>')
                ev.append('</tbody></table></div>')
                new_lines.append('\n'.join(ev))
            current_card_data = None
            current_card_techson = False

        new_lines.append(line)

    return '\n'.join(new_lines)


# ──────────────────────────────────────────────
# Enhance company tab product rows
# ──────────────────────────────────────────────
def enhance_company_product_rows(html_text, patlytics, techson):
    """Add Patlytics score and Techson badge to product rows in company tabs.
    Multiple products can be on a single line, so we use a regex callback
    that processes each product individually."""
    pl_lookup = build_product_patlytics_lookup(patlytics)
    ts_product_map = build_techson_product_set(techson)

    lines = html_text.split('\n')
    new_lines = []
    current_company = None
    badge_count = 0

    for line in lines:
        # Track which company tab we're in
        co_match = re.search(r'id="page-co-([^"]+)"', line)
        if co_match:
            slug = co_match.group(1)
            for co in TARGET_12:
                if co.replace('/', '-') == slug or co == slug:
                    current_company = co
                    break

        # Process lines with product rows (may have many products on one line)
        if current_company and 'class="prod-link"' in line and '<td class="prod-indent">' in line:
            # Split the line into individual <tr> segments and process each
            # Pattern: each product is in a <tr>...<a class="prod-link">NAME</a>...
            #          <span class="r r-xxx">RATING</span></td>...</tr>
            def replace_product_row(m):
                nonlocal badge_count
                tr_html = m.group(0)
                prod_m = re.search(r'class="prod-link">([^<]+)</a>', tr_html)
                if not prod_m:
                    return tr_html
                prod_name = html_mod.unescape(prod_m.group(1))

                pl_data = find_patlytics_for_product(prod_name, current_company, pl_lookup)
                has_techson = (current_company, prod_name) in ts_product_map

                badges = ''
                if pl_data:
                    score_pct = f'{pl_data["best_score"]:.0%}'
                    badges += f' <span class="pi-pat-score" title="Patlytics infringement score">{score_pct}</span>'
                if has_techson:
                    badges += ' <span class="pi-ts-badge" title="Techson also identifies this product">Techson</span>'

                if badges:
                    badge_count += 1
                    tr_html = re.sub(
                        r'(<span class="r r-[^"]*">[^<]*</span>)(</td>)',
                        rf'\1{badges}\2',
                        tr_html,
                        count=1)
                return tr_html

            # Match each <tr> that contains a prod-link (skip div-header-row trs)
            line = re.sub(
                r'<tr><td class="prod-indent">.*?</tr>',
                replace_product_row,
                line)

        new_lines.append(line)

    print(f"  {badge_count} product rows badged in company tabs")
    return '\n'.join(new_lines)


# ──────────────────────────────────────────────
# Standardize company logos across the dashboard
# ──────────────────────────────────────────────
def standardize_company_logos(html_text):
    """Add co-mono icons to sidebar nav items, company tab headers,
    and network table company cells."""
    count = 0

    # 1. Sidebar nav items — add icon before company name
    #    Pattern: <div class="nav-item" data-page="co-Google">Google <span class="b">164</span></div>
    for co in TARGET_12:
        slug = co.replace('/', '-')
        icon = co_icon(co, 16)
        # nav item
        old_nav = f'data-page="co-{slug}">{co} '
        new_nav = f'data-page="co-{slug}">{icon}{co} '
        if old_nav in html_text:
            html_text = html_text.replace(old_nav, new_nav)
            count += 1
        # Company tab header: <h2>Google</h2>
        icon_lg = co_icon(co, 24)
        old_h2 = f'<h2>{co}</h2>'
        new_h2 = f'<h2>{icon_lg} {co}</h2>'
        if old_h2 in html_text:
            html_text = html_text.replace(old_h2, new_h2)
            count += 1

    # 2. Network table cells: <td class="net-co">COMPANY</td>
    for co in TARGET_12:
        icon_sm = co_icon(co, 14)
        old_td = f'<td class="net-co">{co}</td>'
        new_td = f'<td class="net-co">{icon_sm} {co}</td>'
        html_text = html_text.replace(old_td, new_td)

    # 3. Normalize existing co-mono icons from base dashboard (product filter buttons)
    #    The base dashboard uses different colors/letters for some companies.
    replacements = [
        # Apple: ● gray → Ap dark gray
        ('background:#a2aaad;width:16px;height:16px;font-size:9px;line-height:16px">●</span>',
         'background:#555555;width:16px;height:16px;font-size:9px;line-height:16px">Ap</span>'),
        # Meta: blue variant
        ('background:#0668e1;', 'background:#1877F2;'),
        # Microsoft: M → Ms
        ('background:#00a4ef;width:16px;height:16px;font-size:9px;line-height:16px">M</span>',
         'background:#00A4EF;width:16px;height:16px;font-size:9px;line-height:16px">Ms</span>'),
        # Samsung: S → Sa
        ('background:#1428a0;width:16px;height:16px;font-size:9px;line-height:16px">S</span>',
         'background:#1428A0;width:16px;height:16px;font-size:9px;line-height:16px">Sa</span>'),
        # OpenAI: green → purple
        ('background:#10a37f;', 'background:#412991;'),
        # xAI: gray → black
        ('background:#8b8fa3;', 'background:#000000;'),
        # SoftBank/ARM: A → S
        ('background:#ed1c24;width:16px;height:16px;font-size:9px;line-height:16px">A</span>',
         'background:#ED1C24;width:16px;height:16px;font-size:9px;line-height:16px">S</span>'),
    ]
    norm_count = 0
    for old_str, new_str in replacements:
        if old_str in html_text:
            html_text = html_text.replace(old_str, new_str)
            norm_count += 1

    print(f"  {count} nav/header logos added, network table logos injected, {norm_count} old icons normalized")
    return html_text


# ──────────────────────────────────────────────
# Enhance company tabs
# ──────────────────────────────────────────────
def enhance_company_tabs(html_text, patlytics, techson):
    """Insert litigation summary card at top of each company tab."""
    for co in TARGET_12:
        slug = co.replace('/', '-')
        # Try both formats
        for marker_id in [co, slug]:
            marker = f'id="pa-co-{marker_id}"'
            if marker in html_text:
                break
        else:
            continue

        # Aggregate Patlytics data for this company
        pl_scores = []
        for (co_norm, prod), entries in patlytics['by_product'].items():
            if co_norm == co:
                best = max(entries, key=lambda e: e['score'])
                pl_scores.append((prod, best['score'], best['patent_id']))
        pl_scores.sort(key=lambda x: -x[1])

        # Aggregate Techson data
        ts_patents = 0
        ts_revenue = 0
        ts_quality_sum = 0
        for pid, td in techson.items():
            if co in td['target_cos_norm']:
                ts_patents += 1
                ts_revenue += td['revenue']
                ts_quality_sum += td['quality']
        avg_q = ts_quality_sum / ts_patents if ts_patents else 0

        best_pat_score = pl_scores[0][1] if pl_scores else 0

        # Build summary card HTML
        card = []
        card.append(f'<div class="lit-summary-card">')
        card.append(f'<h4>Litigation Assessment <span class="src-badge src-patlytics">Patlytics</span> <span class="src-badge src-techson">Techson</span></h4>')
        card.append(f'<div class="lit-sum-stats">')
        card.append(f'<div class="lit-sum-stat"><div class="lbl">Revenue Risk</div><div class="val" style="color:var(--hi)">{fmt_revenue(ts_revenue)}</div></div>')
        card.append(f'<div class="lit-sum-stat"><div class="lbl">Top Patlytics Score</div><div class="val" style="color:{"var(--cr)" if best_pat_score >= 0.8 else "var(--hi)" if best_pat_score >= 0.5 else "var(--t3)"}">{best_pat_score:.0%}</div></div>')
        card.append(f'<div class="lit-sum-stat"><div class="lbl">Patents Targeting</div><div class="val" style="color:var(--a)">{ts_patents}</div></div>')
        card.append(f'<div class="lit-sum-stat"><div class="lbl">Avg Quality</div><div class="val" style="color:{"var(--gn)" if avg_q >= 6 else "var(--hi)"}">{avg_q:.1f}</div></div>')
        card.append(f'</div>')

        if pl_scores:
            card.append(f'<table class="lit-mini-tbl"><thead><tr><th>Product</th><th>Score</th><th>Source Patent</th></tr></thead><tbody>')
            for prod, sc, pat in pl_scores[:5]:
                cls = score_class(sc)
                card.append(f'<tr><td>{esc(prod)}</td><td class="{cls}">{sc:.0%}</td>'
                            f'<td style="font-size:10px"><a href="#" onclick="goToPatent(\'{pat}\');return false" '
                            f'style="color:var(--a);text-decoration:none">{pat}</a></td></tr>')
            card.append(f'</tbody></table>')

        card.append(f'</div>')
        card_html = '\n'.join(card)

        # Insert before the products table card
        # Find the div with this marker ID (may have style attr after)
        marker_pattern = re.compile(
            r'(<div class="card" ' + re.escape(marker) + r'[^>]*>)')
        m = marker_pattern.search(html_text)
        if m:
            html_text = html_text[:m.start()] + card_html + '\n' + html_text[m.start():]

    return html_text


# ──────────────────────────────────────────────
# Navigation and JS updates
# ──────────────────────────────────────────────
NEW_JS = """
// === LITIGATION DASHBOARD JS ===
function togglePatDetail(patId) {
  var el = document.getElementById('pd-' + patId);
  if (!el) return;
  var isOpen = el.classList.contains('open');
  // Close all
  document.querySelectorAll('.pat-detail.open').forEach(function(d) { d.classList.remove('open'); });
  if (!isOpen) el.classList.add('open');
}

function toggleLit(card) {
  card.classList.toggle('open');
}

function goToCompany(co) {
  var slug = 'co-' + co.replace('/', '-');
  var idx = P.indexOf(slug);
  if (idx >= 0) sp(idx);
}

function goToPatent(patId) {
  sp(P.indexOf('patents'));
  setTimeout(function() {
    var detail = document.getElementById('pd-' + patId);
    if (detail) {
      document.querySelectorAll('.pat-detail.open').forEach(function(d) { d.classList.remove('open'); });
      detail.classList.add('open');
      var c = document.querySelector('.content');
      if (c) c.scrollTop = detail.offsetTop - c.offsetTop - 60;
    }
  }, 60);
}

function togglePatScoreRows(patId, link) {
  var tbl = document.getElementById('pst-' + patId);
  if (!tbl) return;
  var extras = tbl.querySelectorAll('.pst-extra');
  var showing = extras[0] && extras[0].style.display !== 'none';
  extras.forEach(function(tr) { tr.style.display = showing ? 'none' : ''; });
  link.textContent = showing ? 'Show ' + extras.length + ' more' : 'Show fewer';
}

function applyPatFilters() {
  var qf = document.getElementById('pat-q-filter');
  var sf = document.getElementById('pat-s-filter');
  var qVal = qf ? qf.value : 'all';
  var sVal = sf ? sf.value : 'all';

  document.querySelectorAll('.pat-section').forEach(function(section) {
    var rows = section.querySelectorAll('.pat-row');
    var visCount = 0;
    rows.forEach(function(row) {
      var show = true;
      var detail = row.nextElementSibling;
      var patId = row.getAttribute('onclick');
      if (patId) {
        patId = patId.replace("togglePatDetail('", '').replace("')", '');
      }

      // Quality filter
      if (qVal !== 'all' && typeof TS !== 'undefined' && patId) {
        var td = TS[patId];
        if (td) {
          if (qVal === 'high' && td.q < 7) show = false;
          if (qVal === 'mid' && (td.q < 4 || td.q >= 7)) show = false;
          if (qVal === 'low' && td.q >= 4) show = false;
        } else if (qVal !== 'all') {
          show = false;
        }
      }

      // Status filter
      if (sVal !== 'all' && typeof TS !== 'undefined' && patId) {
        var td2 = TS[patId];
        if (td2) {
          if (td2.st.toLowerCase() !== sVal.toLowerCase()) show = false;
        } else if (sVal !== 'all') {
          show = false;
        }
      }

      row.style.display = show ? '' : 'none';
      if (detail && detail.classList.contains('pat-detail')) {
        if (!show) detail.classList.remove('open');
      }
      if (show) visCount++;
    });
  });
}

/* Product card sort */
var _piSortOrder = [];
document.querySelectorAll('.pi').forEach(function(el, i) { _piSortOrder.push({el: el, idx: i}); });

function sortProducts(mode) {
  var container = document.querySelector('.pi') ? document.querySelector('.pi').parentNode : null;
  if (!container) return;
  var cards = Array.from(container.querySelectorAll('.pi'));
  cards.sort(function(a, b) {
    if (mode === 'patlytics') {
      var pa = parseFloat(a.dataset.patlytics || '0');
      var pb = parseFloat(b.dataset.patlytics || '0');
      return pb - pa;
    }
    if (mode === 'techson') {
      var ta = parseInt(a.dataset.techson || '0');
      var tb = parseInt(b.dataset.techson || '0');
      if (tb !== ta) return tb - ta;
      var pa2 = parseFloat(a.dataset.patlytics || '0');
      var pb2 = parseFloat(b.dataset.patlytics || '0');
      return pb2 - pa2;
    }
    // default: original order (use rank-num)
    var ra = parseInt((a.querySelector('.rank-num') || {}).textContent || '999');
    var rb = parseInt((b.querySelector('.rank-num') || {}).textContent || '999');
    return ra - rb;
  });
  cards.forEach(function(c) { container.appendChild(c); });
}
document.querySelectorAll('.pi-sort-btn').forEach(function(btn) {
  btn.addEventListener('click', function() {
    document.querySelectorAll('.pi-sort-btn').forEach(function(b) { b.classList.remove('active'); });
    this.classList.add('active');
    sortProducts(this.dataset.sort);
  });
});
"""


# ──────────────────────────────────────────────
# Main orchestrator
# ──────────────────────────────────────────────
def main():
    print("Loading Patlytics data...")
    patlytics = load_patlytics()
    print(f"  {len(patlytics['by_patent'])} patents, {len(patlytics['by_product'])} products")

    print("Loading Techson data...")
    techson = load_techson()
    print(f"  {len(techson)} patents")

    print("Reading index.html...")
    with open(INPUT_HTML, 'r') as f:
        html_text = f.read()

    # 1. Update title
    html_text = html_text.replace(
        '<title>UltronAI \u2014 Patent Portfolio Dashboard</title>',
        '<title>UltronAI \u2014 Patent Intelligence Dashboard</title>')

    # Remove 'active' class from patents page (litigation page will be default)
    html_text = html_text.replace(
        'class="page active" id="page-patents"',
        'class="page" id="page-patents"')

    # Update sidebar subtitle
    html_text = html_text.replace(
        '<p>Patent Portfolio Dashboard</p>',
        '<p>Patent Intelligence Dashboard</p>')

    # Remove 'active' from the patents nav item (litigation will be default)
    html_text = html_text.replace(
        '<div class="nav-item active" data-page="patents">',
        '<div class="nav-item" data-page="patents">')

    # 2. Reorganize patent categories
    print("Reorganizing patent categories...")
    html_text = reorganize_patents(html_text)

    # 2b. Remap old category cross-references (product card links to patent categories)
    print("Remapping category cross-references...")
    # Cosine Embedding → Neural Network Architecture
    html_text = html_text.replace(
        'data-pat-cat="cosine-embedding-similarity"',
        'data-pat-cat="neural-network-architecture"')
    html_text = html_text.replace(
        '>Cosine Embedding</a>',
        '>Neural Network</a>')
    # Retail / Product AI → Retail: Product Detection (the broader retail category)
    html_text = html_text.replace(
        'data-pat-cat="retail-product-ai"',
        'data-pat-cat="retail-product-detection"')
    html_text = html_text.replace(
        '>Retail/Product AI</a>',
        '>Retail: Product Detection</a>')
    remap_count = 0
    for old in ['cosine-embedding-similarity', 'retail-product-ai']:
        if old in html_text:
            print(f"  WARNING: '{old}' still present after remap")
        else:
            remap_count += 1
    print(f"  Remapped {remap_count}/2 old categories")

    # 3. Inject CSS before </style>
    print("Injecting CSS...")
    html_text = html_text.replace('</style>', f'{NEW_CSS}\n</style>')

    # 4. Build Litigation page
    print("Building Litigation page...")
    lit_page_html = build_litigation_page(patlytics, techson)

    # 5. Build JS data constants
    print("Building JS data constants...")
    # Techson constant (by patent ID)
    ts_js = {}
    for pid, td in techson.items():
        ts_js[pid] = {
            'q': td['quality'], 'rev': td['revenue'],
            'b': td['bundle_no'], 'bt': td['bundle_title'],
            'st': td['status'], 'exp': td['expiration'],
            'cos': list(set(td['target_cos_norm'])),
        }
    ts_json = json.dumps(ts_js, separators=(',', ':'))

    # Patlytics constant (by patent ID — top scores only)
    pl_js = {}
    for pid, entries in patlytics['by_patent'].items():
        sorted_e = sorted(entries, key=lambda e: -e['score'])[:15]
        pl_js[pid] = [{'c': e['co_norm'], 'p': e['prod'],
                        's': e['score'], 'd': e['docs']}
                       for e in sorted_e]
    pl_json = json.dumps(pl_js, separators=(',', ':'))

    # 6. Enhance patent rows
    print("Enhancing patent rows...")
    html_text = enhance_patent_rows(html_text, patlytics, techson)

    # 7. Enhance product cards
    print("Enhancing product cards...")
    html_text = enhance_product_cards(html_text, patlytics, techson)

    # 7b. Add sort buttons to product toolbar
    sort_buttons = ('<div class="pi-sort-wrap"><label>Sort</label>'
                    '<button class="pi-sort-btn active" data-sort="default">Relevance</button>'
                    '<button class="pi-sort-btn" data-sort="patlytics">Patlytics</button>'
                    '<button class="pi-sort-btn" data-sort="techson">Techson</button>'
                    '</div>')
    html_text = html_text.replace(
        '<input class="search-input pi-search"',
        f'{sort_buttons}<input class="search-input pi-search"')

    # 7c. Enhance company tab product rows with Patlytics/Techson badges
    print("Enhancing company tab product rows...")
    html_text = enhance_company_product_rows(html_text, patlytics, techson)

    # 8. Enhance company tabs
    print("Enhancing company tabs...")
    html_text = enhance_company_tabs(html_text, patlytics, techson)

    # 9. Insert Litigation page into HTML
    # Find the first <div class="page" and insert before it
    first_page = re.search(r'<div class="page"', html_text)
    if first_page:
        html_text = html_text[:first_page.start()] + lit_page_html + '\n' + html_text[first_page.start():]

    # 10. Update sidebar navigation — merge Targets + Patents under one Litigation group
    # Remove the standalone Patents nav group
    html_text = re.sub(
        r'<div class="nav-grp ng-pat"><div class="nav-sec">PATENTS</div>\s*'
        r'<div class="nav-item" data-page="patents">[^<]*</div></div>\s*',
        '',
        html_text)
    # Insert combined Litigation group before Products
    nav_lit = ('<div class="nav-grp ng-lit"><div class="nav-sec">Litigation</div>\n'
               '<div class="nav-item" data-page="litigation">Targets</div>\n'
               '<div class="nav-item" data-page="patents">Patents</div>\n</div>\n')
    html_text = html_text.replace(
        '<div class="nav-grp ng-prod">',
        f'{nav_lit}<div class="nav-grp ng-prod">')

    # 10b. Standardize company logos across the dashboard
    print("Standardizing company logos...")
    html_text = standardize_company_logos(html_text)

    # 11. Update JS P array — prepend 'litigation' so page-litigation is found
    html_text = html_text.replace(
        "const P=['patents'",
        "const P=['litigation','patents'")

    # 12. Update JS T dict — add litigation title
    html_text = html_text.replace(
        "const T={'patents'",
        "const T={'litigation':'Litigation Targets','patents'")

    # 13. Inject JS data + functions before the existing const PC
    js_inject = f'\nconst TS = {ts_json};\nconst PL = {pl_json};\n{NEW_JS}\n'
    pc_match = re.search(r'const PC\s*=\s*\{', html_text)
    if pc_match:
        html_text = html_text[:pc_match.start()] + js_inject + html_text[pc_match.start():]

    # 14. Add category overview + legend + filter bar to Patents page
    print("Adding category overview and filters...")
    cat_overview_html = build_category_overview()

    pat_filter_html = """<div class="pat-filters">
<select id="pat-q-filter" class="pat-filter-sel" onchange="applyPatFilters()">
<option value="all">All Quality</option>
<option value="high">High (7-9)</option>
<option value="mid">Medium (4-6)</option>
<option value="low">Low (2-3)</option>
</select>
<select id="pat-s-filter" class="pat-filter-sel" onchange="applyPatFilters()">
<option value="all">All Status</option>
<option value="Active">Active</option>
<option value="Pending">Pending</option>
</select>
</div>"""

    # Insert after the stats grid on patents page, before first pat-section
    patents_page_match = re.search(r'id="page-patents"', html_text)
    if patents_page_match:
        pat_sec_match = re.search(r'<div class="pat-section"', html_text[patents_page_match.start():])
        if pat_sec_match:
            insert_pos = patents_page_match.start() + pat_sec_match.start()
            insert_block = cat_overview_html + '\n' + pat_filter_html + '\n'
            html_text = html_text[:insert_pos] + insert_block + html_text[insert_pos:]

    # 15. Write output
    print(f"Writing {OUTPUT_HTML}...")
    with open(OUTPUT_HTML, 'w') as f:
        f.write(html_text)

    # Stats
    print(f"\nDone! Output: {OUTPUT_HTML}")
    print(f"  File size: {os.path.getsize(OUTPUT_HTML):,} bytes")
    print(f"  Patlytics: {len(patlytics['by_patent'])} patents, {len(patlytics['by_product'])} products")
    print(f"  Techson: {len(techson)} patents")
    total_rev = sum(td['revenue'] for td in techson.values())
    print(f"  Total revenue risk: {fmt_revenue(total_rev)}")


if __name__ == '__main__':
    main()
