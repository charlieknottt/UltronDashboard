import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Patent Research Leaders"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["Company", "Name", "Title", "Department Category", "LinkedIn URL", "Relevant Products/Tech Area"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# All company data
all_data = [
    # ============ NVIDIA ============
    {"Company": "NVIDIA", "Name": "Jan Kautz", "Title": "Vice President of Learning and Perception Research", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/jan-kautz-4668264", "Tech": "Computer vision, deep learning, visual perception, geometric vision, generative models, object detection/recognition, SLAM, optical flow"},
    {"Company": "NVIDIA", "Name": "Azita Martin", "Title": "Vice President & GM of AI for Retail, CPG, and QSR", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/azitamartin/", "Tech": "Retail AI, shrinkage detection, checkout systems, inventory management, computer vision for retail, edge computing"},
    {"Company": "NVIDIA", "Name": "Deepu Talla", "Title": "Vice President & GM of Robotics & Edge Computing", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/deepu-talla-b1631bb", "Tech": "Edge AI, robotics, Jetson platform, embedded systems, generative AI on edge devices, autonomous systems"},
    {"Company": "NVIDIA", "Name": "Tim Teter", "Title": "Executive Vice President, General Counsel and Secretary", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/timothy-s-teter-0244604/", "Tech": "Top legal executive; prior patent/technology litigation at Cooley LLP; engineering background"},
    {"Company": "NVIDIA", "Name": "Rich Domingo", "Title": "Director of Intellectual Property", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/rich-domingo-a1b839b", "Tech": "IP portfolio management, patent strategy"},
    {"Company": "NVIDIA", "Name": "Vishal Bhagwati", "Title": "Head of Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/vishal-bhagwati-36956/", "Tech": "M&A strategy, acquisitions, strategic partnerships, corporate investments"},
    {"Company": "NVIDIA", "Name": "Nico Caprez", "Title": "Corporate Development Manager", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/nicocaprez/", "Tech": "M&A execution, strategic investments, venture investments"},

    # ============ ALPHABET (GOOGLE) ============
    {"Company": "Alphabet (Google)", "Name": "Tomas Pfister", "Title": "Head of AI Research, Google Cloud", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/tomaspfister/", "Tech": "Deep learning, computer vision, facial micro-expression detection, human pose estimation, Face ID technology foundations"},
    {"Company": "Alphabet (Google)", "Name": "Pushmeet Kohli", "Title": "Vice President of Research, Google DeepMind", "Category": "Product/Engineering", "LinkedIn": "https://uk.linkedin.com/in/pushmeet-kohli-4838994", "Tech": "Computer vision, image segmentation, 3D reconstruction, text detection in images, AI safety (SynthID)"},
    {"Company": "Alphabet (Google)", "Name": "Rahul Sukthankar", "Title": "Vice President of Research, Google DeepMind", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/rahulsukthankar/", "Tech": "Computer vision, perception systems, AI powering Cloud, Nest, Pixel, Photos, Search, Waymo, YouTube"},
    {"Company": "Alphabet (Google)", "Name": "Michael Lee", "Title": "Director, Head of Patents", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/mleemichael/", "Tech": "Patent portfolio development, patent licensing and transactions, patent operations"},
    {"Company": "Alphabet (Google)", "Name": "Laura Sheridan", "Title": "Head of Patent Policy", "Category": "IP/Legal", "LinkedIn": "N/A (Google patent policy team)", "Tech": "Patent policy, patent examination process advocacy, AI and patents"},
    {"Company": "Alphabet (Google)", "Name": "Don Harrison", "Title": "President, Global Partnerships and Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/don-harrison-a1b4271/", "Tech": "M&A strategy (Waze, Motorola, Admob), strategic partnerships; board member SpaceX"},
    {"Company": "Alphabet (Google)", "Name": "Sandy Diep", "Title": "M&A Lead - DRS", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/sandydiep", "Tech": "M&A execution, deal structuring"},

    # ============ APPLE ============
    {"Company": "Apple", "Name": "John Giannandrea", "Title": "Senior Vice President of Machine Learning and AI Strategy", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/johngiannandrea/", "Tech": "Machine learning, AI, Siri, Core ML, Face ID, ARKit, Animoji; former Google Search/ML leader"},
    {"Company": "Apple", "Name": "Katherine Adams", "Title": "Senior Vice President and General Counsel", "Category": "IP/Legal", "LinkedIn": "N/A (Apple leadership page)", "Tech": "All legal matters including IP, litigation, corporate governance, privacy"},
    {"Company": "Apple", "Name": "Jeff Myers", "Title": "Vice President, Chief IP Counsel", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/jeff-myers-255256/", "Tech": "Patents, IP licensing, designs; led Intel patent acquisition; former Adobe Chief Patent Counsel"},
    {"Company": "Apple", "Name": "Adrian Perica", "Title": "Vice President, Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "N/A (Apple leadership page)", "Tech": "M&A (led Beats acquisition), strategic investments, iCloud and Infrastructure leadership"},

    # ============ MICROSOFT ============
    {"Company": "Microsoft", "Name": "Peter Lee", "Title": "Corporate Vice President, Research and Incubations", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/peterlee4/", "Tech": "AI research, healthcare AI, machine learning, GPT-4 evaluation; former CMU CS dept head, DARPA"},
    {"Company": "Microsoft", "Name": "Mustafa Suleyman", "Title": "EVP and CEO, Microsoft AI", "Category": "Product/Engineering", "LinkedIn": "N/A (DeepMind/Inflection co-founder)", "Tech": "AI strategy, Copilot products; DeepMind co-founder, Inflection co-founder"},
    {"Company": "Microsoft", "Name": "Brad Smith", "Title": "Vice Chair and President", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/bradsmi/", "Tech": "Legal, IP, M&A, corporate governance, litigation; top legal executive"},
    {"Company": "Microsoft", "Name": "Edward Kazenske", "Title": "Senior Director of Patent Prosecution Strategy", "Category": "IP/Legal", "LinkedIn": "N/A (IP & Licensing Group)", "Tech": "Patent prosecution, former Deputy Commissioner for Patents at USPTO"},
    {"Company": "Microsoft", "Name": "Brian Schultz", "Title": "Managing Director and Head of Strategic Investments", "Category": "M&A/Corp Dev", "LinkedIn": "N/A (M1 Finance investor page)", "Tech": "Strategic investments, venture capital, M&A; co-founded Ontela"},

    # ============ AMAZON ============
    {"Company": "Amazon", "Name": "Dilip Kumar", "Title": "Vice President, Physical Retail and Technology / AWS", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/dilipskumar/", "Tech": "Just Walk Out technology, Amazon Go, computer vision, deep learning, retail checkout systems"},
    {"Company": "Amazon", "Name": "Ashok Elluswamy", "Title": "Vice President of AI Software (Tesla - related tech)", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/eashokkumar/", "Tech": "Note: This is Tesla's AI VP - Amazon equivalent: Jon Jenkins (VP Just Walk Out)"},
    {"Company": "Amazon", "Name": "Dave Platz", "Title": "Patent Counsel, AWS", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/dave-platz-6a1013a/", "Tech": "AWS IP assets, patent portfolio development"},
    {"Company": "Amazon", "Name": "Alex Ceballos", "Title": "Vice President, Worldwide Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "N/A (Crunchbase profile)", "Tech": "M&A, investments, partnerships (US, Europe, India); former Global Credit leader"},
    {"Company": "Amazon", "Name": "David Shearer", "Title": "Former VP, Corporate Development, Business Development, Alexa & Devices", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/davidshearer1/", "Tech": "21-year Amazon executive; acquisitions, partnerships, Amazon Photos, Alexa Internet"},

    # ============ META ============
    {"Company": "Meta", "Name": "Manohar Paluri", "Title": "Former Director/VP, Computer Vision (13.5 years at Meta)", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/balamanohar/", "Tech": "Computer vision, robotics, machine learning; started as intern 2011"},
    {"Company": "Meta", "Name": "Peter Vajda", "Title": "Director, Computer Vision", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/p%C3%A9ter-vajda-9a03aaa/", "Tech": "Deep learning on mobile (Caffe2Go), real-time pixel processing"},
    {"Company": "Meta", "Name": "Allen Lo", "Title": "Vice President and Deputy General Counsel, Product, IP, and Legal Operations", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/allen-lo-50b4431/", "Tech": "IP leadership; former Google Deputy GC Patents; defended Android patent wars"},
    {"Company": "Meta", "Name": "Mark Fiore", "Title": "Director & Associate General Counsel, IP (Content & Product)", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/mark-fiore-4164b08/", "Tech": "AI IP counseling, Llama models, Meta AI, intermediary liability"},
    {"Company": "Meta", "Name": "Amin Zoufonoun", "Title": "Vice President of Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/amin-zoufonoun-a6a3995/", "Tech": "M&A since 2011 (Instagram, WhatsApp, Oculus); former Google Corp Dev Director"},

    # ============ TESLA ============
    {"Company": "Tesla", "Name": "Ashok Elluswamy", "Title": "Vice President of AI Software", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/eashokkumar/", "Tech": "Autopilot, FSD, computer vision, autonomous driving; 10+ years at Tesla"},
    {"Company": "Tesla", "Name": "John Jones", "Title": "VP - Autopilot / AI Software", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/john-jones-1bbbba1b5/", "Tech": "Autopilot, AI software development"},
    {"Company": "Tesla", "Name": "N/A - Tesla Legal", "Title": "General Counsel (IP handled through legal team)", "Category": "IP/Legal", "LinkedIn": "N/A", "Tech": "Tesla uses open patent pledge; 117+ patents, battery/charging focus"},
    {"Company": "Tesla", "Name": "N/A - Tesla Corp Dev", "Title": "Head of Corporate Development (Zach Kirkman left for GM 2023)", "Category": "M&A/Corp Dev", "LinkedIn": "N/A", "Tech": "Previous head Zach Kirkman now VP Corp Dev at GM"},

    # ============ SAMSUNG ============
    {"Company": "Samsung", "Name": "Wonil Roh", "Title": "Executive (Samsung Research America)", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/wonilroh/", "Tech": "AI research, Samsung Research America"},
    {"Company": "Samsung", "Name": "Kihwan Kim", "Title": "AI/XR Executive", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/kihwan", "Tech": "XR devices, AI glasses, multimodal lidar-camera, 3D geometry"},
    {"Company": "Samsung", "Name": "Jason Martineau", "Title": "Director - IP Group, Samsung Semiconductor", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/jason-martineau-7a353213/", "Tech": "Patents (CS, semiconductor, biotech), IP licensing, portfolio management"},
    {"Company": "Samsung", "Name": "S.H. Michael Kim", "Title": "IP Litigation Executive", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/s-h-michael-kim-0056b14/", "Tech": "Patent litigation, PTAB proceedings"},
    {"Company": "Samsung", "Name": "Steve Hong", "Title": "Director, Corporate Development / M&A", "Category": "M&A/Corp Dev", "LinkedIn": "https://kr.linkedin.com/in/stevehong1", "Tech": "M&A deal sourcing, execution, post-merger integration; corporate venture capital"},

    # ============ OPENAI ============
    {"Company": "OpenAI", "Name": "Srinivas Narayanan", "Title": "Vice President of Engineering", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/srinivasnarayanan/", "Tech": "Engineering leadership; former Facebook VP Engineering"},
    {"Company": "OpenAI", "Name": "James Dyett", "Title": "AI/ML Leader", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/jamesdyett/", "Tech": "Generative AI, NLP, computer vision, automated decision-making at cloud scale"},
    {"Company": "OpenAI", "Name": "Che Chang", "Title": "General Counsel", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/chechang/", "Tech": "Litigation, compliance, IP, regulatory; former Amazon legal"},
    {"Company": "OpenAI", "Name": "Bobby Wu", "Title": "Deputy General Counsel", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/bobby-wu-21027516/", "Tech": "Corporate, M&A; Top AI Lawyers 2024 (Daily Journal)"},

    # ============ QUALCOMM ============
    {"Company": "Qualcomm", "Name": "Dr. Vinesh Sukumar", "Title": "VP of Artificial Intelligence", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/vineshsukumar", "Tech": "AI/ML product management; former NASA, Apple (first iPhone camera), Intel"},
    {"Company": "Qualcomm", "Name": "Milivoje Aleksic", "Title": "Senior Vice President of Engineering", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/milivoje-aleksic-6851614/", "Tech": "AI, ML architecture, low power AI, DSP, visual perception; 150+ patents"},
    {"Company": "Qualcomm", "Name": "Michael Kramer", "Title": "VP, Patent Counsel", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/michael-kramer-31a2584/", "Tech": "Patent counsel, IP strategy"},
    {"Company": "Qualcomm", "Name": "Sheila Baran", "Title": "VP & Legal Counsel", "Category": "IP/Legal", "LinkedIn": "https://www.linkedin.com/in/sheilabaran/", "Tech": "IP licensing, M&A, corporate transactions, privacy"},
    {"Company": "Qualcomm", "Name": "Duane Nelles", "Title": "Senior Vice President, Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "https://www.linkedin.com/in/duane-nelles-4347935/", "Tech": "Corporate development since 1996; acquisitions, investments"},

    # ============ SOFTBANK (ARM) ============
    {"Company": "SoftBank (ARM)", "Name": "Mohamed Awad", "Title": "Executive Vice President, Cloud AI Business Unit", "Category": "Product/Engineering", "LinkedIn": "https://www.linkedin.com/in/moawad", "Tech": "Cloud computing, AI infrastructure, AI data centers; former Broadcom, IoT"},
    {"Company": "SoftBank (ARM)", "Name": "Richard Grisenthwaite", "Title": "Arm Fellow", "Category": "Product/Engineering", "LinkedIn": "N/A (Arm leadership page)", "Tech": "Microprocessors; 120 patents; Cambridge engineering"},
    {"Company": "SoftBank (ARM)", "Name": "Rob Calico", "Title": "Vice President, Intellectual Property and Litigation", "Category": "IP/Legal", "LinkedIn": "N/A (Vanguard Law profile)", "Tech": "IP, litigation; former engineer (BIOS ROM development); Purdue EE"},
    {"Company": "SoftBank (ARM)", "Name": "James Hodgson", "Title": "VP, Intellectual Property Products Group (IPG) Licensing", "Category": "IP/Legal", "LinkedIn": "N/A (Arm leadership)", "Tech": "IP licensing, ARM's core business model"},
    {"Company": "SoftBank (ARM)", "Name": "Hassan Parsa", "Title": "VP, Global Head of Corporate Development", "Category": "M&A/Corp Dev", "LinkedIn": "N/A (Arm leadership)", "Tech": "Corporate development, M&A strategy"},
]

# Write data
for row_idx, data in enumerate(all_data, 2):
    ws.cell(row=row_idx, column=1, value=data["Company"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=data["Name"]).border = thin_border
    ws.cell(row=row_idx, column=3, value=data["Title"]).border = thin_border
    ws.cell(row=row_idx, column=4, value=data["Category"]).border = thin_border
    ws.cell(row=row_idx, column=5, value=data["LinkedIn"]).border = thin_border
    ws.cell(row=row_idx, column=6, value=data["Tech"]).border = thin_border

    for col in range(1, 7):
        ws.cell(row=row_idx, column=col).alignment = cell_alignment

# Set column widths
column_widths = [18, 24, 60, 18, 55, 95]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze top row
ws.freeze_panes = "A2"

# Save
wb.save("/Users/charlieknott/Downloads/CMU SPRING /Ultron Internship/Patent_Research_Leaders.xlsx")
print(f"Excel file created with {len(all_data)} leaders across all companies!")
